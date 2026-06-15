from __future__ import annotations

from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
import re
from typing import Iterable

import openpyxl


PRIMARY_DIMENSIONS = {
  "org_model": "ORG_MODEL(PRODUCT_DESC)",
  "segment": "Segment",
  "odm_oem": "ODM_OEM",
}

BREAKDOWN_DIMENSIONS = {
  "muc_module": "MUC_MODULE",
  "problem_mapping": "PROBLEM_Mapping",
}

FILTER_FIELDS = {
  "model": "ORG_MODEL(PRODUCT_DESC)",
  "segment": "Segment",
  "odm_oem": "ODM_OEM",
  "muc_module": "MUC_MODULE",
  "problem_mapping": "PROBLEM_Mapping",
}

RAW_SHEET_NAME = "raw data"
SUMMARY_SHEET_NAME = "SUMMARY_IEC"
MAX_CONSECUTIVE_EMPTY_ROWS = 200

COLUMN_ALIASES = {
  "RMA_NO": ("RMA_NO", "RMA NO", "RMA_NUMBER", "RMA NUMBER"),
  "Week": ("Week", "WK", "WEEK_NO", "WEEK NO"),
  "ORG_MODEL(PRODUCT_DESC)": (
    "ORG_MODEL(PRODUCT_DESC)",
    "ORG MODEL(PRODUCT DESC)",
    "ORG_MODEL",
    "MODEL",
    "PRODUCT_DESC",
    "PRODUCT DESC",
  ),
  "Segment": ("Segment", "SEGMENT"),
  "ODM_OEM": ("ODM_OEM", "ODM/OEM", "ODM OEM"),
  "MUC_MODULE": ("MUC_MODULE", "MUC MODULE", "MODULE"),
  "PROBLEM_Mapping": ("PROBLEM_Mapping", "PROBLEM Mapping", "Problem Mapping", "PROBLEM_MAPPING"),
}


@dataclass(frozen=True)
class WorkbookUpload:
  path: Path
  filename: str


def _clean(value) -> str:
  if value is None:
    return ""
  text = str(value).replace("\xa0", " ").strip()
  return re.sub(r"\s+", " ", text)


def _normalized_header(value) -> str:
  return re.sub(r"[^A-Z0-9]", "", _clean(value).upper())


def _sheet_name(workbook, expected_name: str) -> str | None:
  expected_key = _normalized_header(expected_name)
  for sheet_name in workbook.sheetnames:
    if _normalized_header(sheet_name) == expected_key:
      return sheet_name
  return None


def _header_index(headers: list[str], candidates: Iterable[str]) -> int | None:
  normalized_candidates = {_normalized_header(candidate) for candidate in candidates}
  for index, header in enumerate(headers):
    if _normalized_header(header) in normalized_candidates:
      return index
  return None


def _header_index_by_pattern(headers: list[str], patterns: Iterable[str]) -> int | None:
  for index, header in enumerate(headers):
    clean_header = _clean(header)
    for pattern in patterns:
      if re.search(pattern, clean_header, flags=re.IGNORECASE):
        return index
  return None


def _detect_source_type(filename: str) -> str:
  upper_name = filename.upper()
  if "GAMING" in upper_name:
    return "Gaming NB"
  if re.search(r"(^|[^A-Z])PC[-_ ]?NB", upper_name):
    return "PC NB"
  return "Uploaded"


def _week_sort_key(week: str) -> tuple[int, int, str]:
  match = re.search(r"W?(\d{2})(\d{2})", week or "")
  if not match:
    return (0, 0, week or "")
  return (int(match.group(1)), int(match.group(2)), week)


def _number(value) -> float | None:
  if value is None or value == "":
    return None
  if isinstance(value, (int, float)):
    return float(value)
  text = _clean(value).replace(",", "")
  if not text or text in {"-", "N/A", "#N/A"}:
    return None
  try:
    return float(text)
  except ValueError:
    return None


def _percent_decimal(value) -> float | None:
  if value is None or value == "":
    return None
  if isinstance(value, (int, float)):
    number = float(value)
    return number / 100 if number > 1 else number
  text = _clean(value)
  if not text or text in {"-", "N/A", "#N/A"}:
    return None
  is_percent = text.endswith("%")
  number = _number(text.rstrip("%"))
  if number is None:
    return None
  return number / 100 if is_percent or number > 1 else number


def _parse_summary_iec(workbook, filename: str) -> dict[str, dict]:
  summary_sheet_name = _sheet_name(workbook, SUMMARY_SHEET_NAME)
  if not summary_sheet_name:
    return {}

  worksheet = workbook[summary_sheet_name]
  header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
  headers = [_clean(value) for value in header_row]
  model_index = _header_index(headers, ("MODEL", "Model"))
  failure_index = _header_index(headers, ("IW Failure Q'ty", "IW Failure Qty", "IW Failure Quantity"))
  cfr_index = _header_index_by_pattern(
    headers,
    (
      r"^20\d{2}\s+CFR\(A\)\s+for\s+model$",
      r"CFR\(A\).*for\s+model",
    ),
  )
  target_index = _header_index_by_pattern(
    headers,
    (
      r"^20\d{2}\s+Target$",
      r"\bTarget\b",
    ),
  )

  if model_index is None or failure_index is None or cfr_index is None:
    return {}

  summary: dict[str, dict] = {}
  for row in worksheet.iter_rows(min_row=2, values_only=True):
    if not any(row):
      continue
    model = _clean(row[model_index]) if model_index < len(row) else ""
    if not model:
      continue
    failure_qty = _number(row[failure_index] if failure_index < len(row) else None)
    cfr_decimal = _percent_decimal(row[cfr_index] if cfr_index < len(row) else None)
    target_decimal = _percent_decimal(row[target_index] if target_index is not None and target_index < len(row) else None)
    derived_act = None
    if failure_qty is not None and cfr_decimal and cfr_decimal > 0:
      derived_act = failure_qty / cfr_decimal

    summary[model] = {
      "model": model,
      "source_file": filename,
      "source_type": _detect_source_type(filename),
      "summary_failure_qty": failure_qty,
      "summary_cfr": cfr_decimal,
      "target_cfr": target_decimal,
      "derived_act": derived_act,
    }
  return summary


def parse_workbooks(workbooks: Iterable[WorkbookUpload]) -> dict:
  records: list[dict] = []
  file_summaries: list[dict] = []
  missing_columns_by_file: dict[str, list[str]] = {}
  summary_by_model: dict[str, dict] = {}

  required_columns = {
    "RMA_NO",
    "Week",
    *PRIMARY_DIMENSIONS.values(),
    *BREAKDOWN_DIMENSIONS.values(),
  }

  for upload in workbooks:
    workbook = openpyxl.load_workbook(upload.path, read_only=True, data_only=True)
    try:
      for model, summary in _parse_summary_iec(workbook, upload.filename).items():
        existing = summary_by_model.get(model)
        if existing and existing.get("derived_act") and summary.get("derived_act"):
          existing["derived_act"] += summary["derived_act"]
          existing["summary_failure_qty"] = (existing.get("summary_failure_qty") or 0) + (summary.get("summary_failure_qty") or 0)
          existing["summary_cfr"] = (
            existing["summary_failure_qty"] / existing["derived_act"]
            if existing["derived_act"]
            else None
          )
          continue
        summary_by_model[model] = summary

      raw_sheet_name = _sheet_name(workbook, RAW_SHEET_NAME)
      if not raw_sheet_name:
        missing_columns_by_file[upload.filename] = [RAW_SHEET_NAME]
        file_summaries.append(
          {
            "filename": upload.filename,
            "source_type": _detect_source_type(upload.filename),
            "rows": 0,
            "sheets": workbook.sheetnames,
          }
        )
        continue

      worksheet = workbook[raw_sheet_name]
      header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
      if not header_row:
        missing_columns_by_file[upload.filename] = ["raw data header row"]
        file_summaries.append(
          {
            "filename": upload.filename,
            "source_type": _detect_source_type(upload.filename),
            "rows": 0,
            "sheets": workbook.sheetnames,
          }
        )
        continue
      headers = [_clean(value) for value in header_row]
      index_by_required_field = {
        field_name: _header_index(headers, COLUMN_ALIASES.get(field_name, (field_name,)))
        for field_name in required_columns
      }
      missing = sorted(field_name for field_name, index in index_by_required_field.items() if index is None)
      if missing:
        missing_columns_by_file[upload.filename] = missing
      optional_indices = {
        optional_field: _header_index(headers, (optional_field,))
        for optional_field in ("REGION", "COUNTRY", "PRODUCT_LINE", "CUSTOMER_NAME", "REPAIR_LEVEL")
      }

      row_count = 0
      consecutive_empty_rows = 0
      data_started = False
      for row in worksheet.iter_rows(min_row=2, values_only=True):
        if not any(row):
          if data_started:
            consecutive_empty_rows += 1
            if consecutive_empty_rows >= MAX_CONSECUTIVE_EMPTY_ROWS:
              break
          continue

        data_started = True
        consecutive_empty_rows = 0
        record = {
          "source_file": upload.filename,
          "source_type": _detect_source_type(upload.filename),
        }
        for field_name in required_columns:
          index = index_by_required_field.get(field_name)
          record[field_name] = _clean(row[index]) if index is not None and index < len(row) else ""

        for optional_field, index in optional_indices.items():
          record[optional_field] = _clean(row[index]) if index is not None and index < len(row) else ""

        records.append(record)
        row_count += 1

      file_summaries.append(
        {
          "filename": upload.filename,
          "source_type": _detect_source_type(upload.filename),
          "rows": row_count,
          "sheets": workbook.sheetnames,
        }
      )
    finally:
      workbook.close()

  return {
    "records": records,
    "files": file_summaries,
    "missing_columns": missing_columns_by_file,
    "summary_by_model": summary_by_model,
  }


def dataset_options(records: list[dict], filters: dict | None = None) -> dict[str, list[str]]:
  filters = filters or {}
  options: dict[str, list[str]] = {}
  for option_key, column_name in FILTER_FIELDS.items():
    related_filters = {
      key: value
      for key, value in filters.items()
      if key != option_key
    }
    related_records = _apply_filters(records, related_filters)
    counter = Counter(record.get(column_name, "") or "(blank)" for record in related_records)
    selected = _selected_values(filters.get(option_key, ""))
    for value in selected:
      if value not in counter:
        counter[value] = 0
    options[option_key] = [value for value, _ in counter.most_common()]
  return options


def _selected_values(value) -> set[str]:
  if value is None:
    return set()
  if isinstance(value, str):
    values = [value]
  else:
    values = list(value)
  return {
    str(item).strip()
    for item in values
    if str(item).strip()
  }


def _apply_filters(records: list[dict], filters: dict) -> list[dict]:
  filtered = records
  for filter_key, column_name in FILTER_FIELDS.items():
    selected = _selected_values(filters.get(filter_key, ""))
    if not selected:
      continue
    filtered = [
      record
      for record in filtered
      if (record.get(column_name) or "(blank)") in selected
    ]
  return filtered


def _counter_rows(counter: Counter, limit: int | None = None) -> list[dict]:
  total = sum(counter.values()) or 1
  rows = []
  for label, count in counter.most_common(limit):
    rows.append(
      {
        "label": label or "(blank)",
        "count": count,
        "share": count / total * 100,
      }
    )
  return rows


def analyze_dataset(
  records: list[dict],
  summary_by_model: dict[str, dict] | None = None,
  primary_dimension: str = "org_model",
  breakdown_dimension: str = "problem_mapping",
  filters: dict | None = None,
) -> dict:
  filters = filters or {}
  summary_by_model = summary_by_model or {}
  primary_column = PRIMARY_DIMENSIONS.get(primary_dimension, PRIMARY_DIMENSIONS["org_model"])
  breakdown_column = BREAKDOWN_DIMENSIONS.get(breakdown_dimension, BREAKDOWN_DIMENSIONS["problem_mapping"])
  filtered_records = _apply_filters(records, filters)

  week_counter = Counter(record.get("Week", "") or "(blank)" for record in filtered_records)
  sorted_weeks = sorted(week_counter, key=_week_sort_key)
  latest_week = sorted_weeks[-1] if sorted_weeks else ""
  previous_week = sorted_weeks[-2] if len(sorted_weeks) >= 2 else ""
  latest_count = week_counter.get(latest_week, 0)
  previous_count = week_counter.get(previous_week, 0)

  unique_rma = {
    record.get("RMA_NO")
    for record in filtered_records
    if record.get("RMA_NO")
  }

  primary_counter = Counter(record.get(primary_column, "") or "(blank)" for record in filtered_records)
  breakdown_counter = Counter(record.get(breakdown_column, "") or "(blank)" for record in filtered_records)
  module_counter = Counter(record.get("MUC_MODULE", "") or "(blank)" for record in filtered_records)
  source_counter = Counter(record.get("source_type", "") or "Uploaded" for record in filtered_records)
  filtered_models = {
    record.get("ORG_MODEL(PRODUCT_DESC)", "")
    for record in filtered_records
    if record.get("ORG_MODEL(PRODUCT_DESC)", "")
  }
  act_rows = [
    summary_by_model[model]
    for model in filtered_models
    if model in summary_by_model and summary_by_model[model].get("derived_act")
  ]
  derived_act = sum(row["derived_act"] for row in act_rows)
  filtered_failure_qty = len(filtered_records)
  filtered_cfr = filtered_failure_qty / derived_act if derived_act else None
  target_values = [
    row["target_cfr"]
    for row in act_rows
    if row.get("target_cfr") is not None
  ]
  target_cfr = sum(target_values) / len(target_values) if target_values else None

  cumulative = 0
  total_breakdown = sum(breakdown_counter.values()) or 1
  pareto_rows = []
  for label, count in breakdown_counter.most_common(15):
    cumulative += count
    pareto_rows.append(
      {
        "label": label or "(blank)",
        "count": count,
        "share": count / total_breakdown * 100,
        "cumulative": cumulative / total_breakdown * 100,
      }
    )

  top_primary_labels = [label for label, _ in primary_counter.most_common(8)]
  top_breakdown_labels = [label for label, _ in breakdown_counter.most_common(8)]
  matrix_counts: dict[tuple[str, str], int] = defaultdict(int)
  combo_counter = Counter()
  for record in filtered_records:
    primary_label = record.get(primary_column, "") or "(blank)"
    breakdown_label = record.get(breakdown_column, "") or "(blank)"
    matrix_counts[(primary_label, breakdown_label)] += 1
    combo_counter[(primary_label, breakdown_label)] += 1

  heatmap = []
  for primary_label in top_primary_labels:
    primary_act = None
    if primary_column == "ORG_MODEL(PRODUCT_DESC)":
      primary_summary = summary_by_model.get(primary_label)
      if primary_summary:
        primary_act = primary_summary.get("derived_act")

    values = []
    for breakdown_label in top_breakdown_labels:
      count = matrix_counts[(primary_label, breakdown_label)]
      cfr = count / primary_act if primary_act else None
      values.append(
        {
          "count": count,
          "cfr": cfr,
        }
      )

    heatmap.append(
      {
        "label": primary_label,
        "values": values,
      }
    )

  top_combinations = [
    {
      "primary": primary_label,
      "breakdown": breakdown_label,
      "count": count,
    }
    for (primary_label, breakdown_label), count in combo_counter.most_common(12)
  ]

  trend = [
    {
      "week": week,
      "count": week_counter[week],
    }
    for week in sorted_weeks
  ]

  return {
    "primary_dimension": primary_dimension,
    "primary_label": primary_column,
    "breakdown_dimension": breakdown_dimension,
    "breakdown_label": breakdown_column,
    "filters": filters,
    "record_count": len(filtered_records),
    "total_record_count": len(records),
    "kpis": {
      "records": len(filtered_records),
      "unique_rma": len(unique_rma),
      "week_count": len(sorted_weeks),
      "latest_week": latest_week or "N/A",
      "latest_count": latest_count,
      "previous_week": previous_week or "N/A",
      "previous_count": previous_count,
      "week_delta": latest_count - previous_count,
      "filtered_cfr": filtered_cfr,
      "filtered_failure_qty": filtered_failure_qty,
      "derived_act": derived_act if derived_act else None,
      "act_model_count": len(act_rows),
      "target_cfr": target_cfr,
    },
    "source_mix": _counter_rows(source_counter),
    "trend": trend,
    "share": _counter_rows(primary_counter, 12),
    "module_ratio": _counter_rows(module_counter, 12),
    "pareto": pareto_rows,
    "heatmap_columns": top_breakdown_labels,
    "heatmap": heatmap,
    "top_combinations": top_combinations,
  }
