from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
import pandas as pd


MONTH_NAME_TO_NUMBER = {
  "jan": 1,
  "feb": 2,
  "mar": 3,
  "apr": 4,
  "may": 5,
  "jun": 6,
  "jul": 7,
  "aug": 8,
  "sep": 9,
  "oct": 10,
  "nov": 11,
  "dec": 12,
}

MONTH_SHORT = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def analyze_mail_summary(workbook_path: str | Path) -> dict[str, Any]:
  path = Path(workbook_path)
  asus = _load_asus_source(path)
  hp = _load_hp_source(path, asus["latest_year"], asus["latest_month"])
  totals = _load_total_customer_sheet(path)

  current_year = asus["latest_year"]
  current_month = asus["latest_month"]
  previous_year = current_year - 1
  current_month_label = f"{MONTH_SHORT[current_month - 1]}-{str(current_year)[-2:]}"
  previous_month_label = _format_month_label(*_previous_month(current_year, current_month))
  same_month_last_year_label = f"{MONTH_SHORT[current_month - 1]}-{str(previous_year)[-2:]}"
  ytd_current_label = f"Jan'{str(current_year)[-2:]} ~ {MONTH_SHORT[current_month - 1]}'{str(current_year)[-2:]}"
  ytd_previous_label = f"Jan'{str(previous_year)[-2:]} ~ {MONTH_SHORT[current_month - 1]}'{str(previous_year)[-2:]}"

  current_rows = [
    {
      "label": "HP_AIO (BPS)",
      "previous": hp["bps_previous_month"],
      "current": hp["bps_current_month"],
    },
    {
      "label": "HP_AIO (CPS)",
      "previous": hp["cps_previous_month"],
      "current": hp["cps_current_month"],
    },
    {
      "label": "ASUS",
      "previous": totals["asus_previous_month"],
      "current": totals["asus_current_month"],
    },
  ]
  current_table = _with_deltas(current_rows)
  current_total_previous = sum(item["previous"] for item in current_rows)
  current_total_current = sum(item["current"] for item in current_rows)
  current_total_delta = current_total_current - current_total_previous
  current_total_ratio = _safe_pct(current_total_delta, current_total_previous)

  ytd_rows = [
    {
      "label": "HP_AIO (BPS)",
      "previous": hp["bps_previous_ytd"],
      "current": hp["bps_current_ytd"],
    },
    {
      "label": "HP_AIO (CPS)",
      "previous": hp["cps_previous_ytd"],
      "current": hp["cps_current_ytd"],
    },
    {
      "label": "ASUS",
      "previous": totals["asus_previous_ytd"],
      "current": totals["asus_current_ytd"],
    },
  ]
  ytd_table = _with_deltas(ytd_rows)
  ytd_total_previous = sum(item["previous"] for item in ytd_rows)
  ytd_total_current = sum(item["current"] for item in ytd_rows)
  ytd_total_delta = ytd_total_current - ytd_total_previous
  ytd_total_ratio = _safe_pct(ytd_total_delta, ytd_total_previous)

  hp_current_month_total = totals["hp_current_month"]
  hp_previous_year_same_month_total = totals["hp_same_month_last_year"]
  hp_same_month_delta = hp_current_month_total - hp_previous_year_same_month_total
  asus_same_month_delta = totals["asus_current_month"] - totals["asus_same_month_last_year"]
  overall_same_month_previous = totals["overall_same_month_last_year"]
  overall_same_month_current = totals["overall_current_month"]
  overall_same_month_delta = overall_same_month_current - overall_same_month_previous
  overall_same_month_ratio = _safe_pct(overall_same_month_delta, overall_same_month_previous)

  headline_bullets = [
    (
      f"{current_year}年{current_month}月出貨數較{current_year}年{current_month - 1 if current_month > 1 else 12}月，"
      f"總出貨成長{_format_pct(current_total_ratio)}"
    ),
    (
      f"{ytd_current_label}與{ytd_previous_label}相比，出貨{'增加' if ytd_total_delta >= 0 else '減少'}"
      f"{abs(ytd_total_delta):,}台，{'成長' if (ytd_total_ratio or 0) >= 0 else '衰退'}{abs(ytd_total_ratio or 0):.2f}%"
    ),
    (
      f"{current_year}年{current_month}月ASUS & HP AIO較{previous_year}年{current_month}月比較，"
      f"ASUS {'減少' if asus_same_month_delta < 0 else '增加'}{abs(asus_same_month_delta):,}台，"
      f"HP AIO {'減少' if hp_same_month_delta < 0 else '增加'}{abs(hp_same_month_delta):,}台"
    ),
    (
      f"ASUS Gaming NB & PC 出貨占比："
      f"{current_year}年{current_month}月 Gaming NB {asus['current_month_category_share'].get('Gaming', 0.0):.0f}% / "
      f"PC {asus['current_month_category_share'].get('PC', 0.0):.0f}%"
    ),
    (
      f"{current_year}年累計出貨占比："
      f"Gaming NB {asus['current_ytd_category_share'].get('Gaming', 0.0):.0f}% / "
      f"PC {asus['current_ytd_category_share'].get('PC', 0.0):.0f}%"
    ),
  ]

  detail_bullets = [
    (
      f"{current_year}年{current_month}月出貨數較{previous_month_label}，"
      f"{'成長' if (current_total_ratio or 0) >= 0 else '衰退'}{abs(current_total_ratio or 0):.2f}%"
    ),
    (
      f"{ytd_current_label}累計出貨相較{ytd_previous_label}，"
      f"出貨{'增加' if ytd_total_delta >= 0 else '減少'}{abs(ytd_total_delta):,}台，"
      f"{'成長' if (ytd_total_ratio or 0) >= 0 else '衰退'}{abs(ytd_total_ratio or 0):.2f}%"
    ),
    (
      f"{current_month_label}較{same_month_last_year_label}："
      f"總出貨{_format_signed(overall_same_month_delta)}台，"
      f"ASUS{_format_signed(asus_same_month_delta)}台，"
      f"HP AIO{_format_signed(hp_same_month_delta)}台"
    ),
    (
      f"ASUS {current_month_label} ODM models share approximately "
      f"{asus['current_month_type_share'].get('ODM', 0.0):.2f}%"
    ),
  ]

  source_notes = [
    "Customer-level totals are aligned with the workbook Total sheet because it matches the approved mail summary figures.",
    "Gaming / PC mix is derived from the ASUS source sheet.",
    "Territory mix is not generated in this version because region data is not present in HP-AIO and ASUS source sheets.",
  ]

  return {
    "workbook_name": path.name,
    "current_month_label": current_month_label,
    "previous_month_label": previous_month_label,
    "same_month_last_year_label": same_month_last_year_label,
    "ytd_current_label": ytd_current_label,
    "ytd_previous_label": ytd_previous_label,
    "headline_bullets": headline_bullets,
    "detail_bullets": detail_bullets,
    "current_table": current_table,
    "current_total": {
      "previous": current_total_previous,
      "current": current_total_current,
      "delta": current_total_delta,
      "ratio": current_total_ratio,
    },
    "ytd_table": ytd_table,
    "ytd_total": {
      "previous": ytd_total_previous,
      "current": ytd_total_current,
      "delta": ytd_total_delta,
      "ratio": ytd_total_ratio,
    },
    "gaming_pc_summary": {
      "current_month": asus["current_month_category_share"],
      "current_ytd": asus["current_ytd_category_share"],
    },
    "type_summary": {
        "current_month": asus["current_month_type_share"],
    },
    "source_notes": source_notes,
  }


def _load_asus_source(path: Path) -> dict[str, Any]:
  df = pd.read_excel(path, sheet_name="ASUS")
  df["Model"] = df["Model"].astype(str).str.strip()
  df["Type"] = df.get("Type", "").astype(str).str.strip()
  df["Category"] = df.get("Category", "").astype(str).str.strip()

  date_columns = [col for col in df.columns if isinstance(col, datetime)]
  date_columns.sort()
  month_column_map = {(col.year, col.month): col for col in date_columns}
  latest_col = max(date_columns)
  latest_year = latest_col.year
  latest_month = latest_col.month
  previous_year, previous_month = _previous_month(latest_year, latest_month)

  total_row = df.loc[df["Model"].str.casefold() == "total"].iloc[0]
  current_ytd_columns = [month_column_map[(latest_year, month)] for month in range(1, latest_month + 1)]

  detail_rows = df.loc[df["Model"].str.casefold() != "total"].copy()

  current_month_category_share = _group_share(detail_rows, "Category", month_column_map[(latest_year, latest_month)])
  current_ytd_category_share = _group_share(detail_rows, "Category", current_ytd_columns)
  current_month_type_share = _group_share(detail_rows, "Type", month_column_map[(latest_year, latest_month)])

  return {
    "latest_year": latest_year,
    "latest_month": latest_month,
    "current_month_category_share": current_month_category_share,
    "current_ytd_category_share": current_ytd_category_share,
    "current_month_type_share": current_month_type_share,
  }


def _load_total_customer_sheet(path: Path) -> dict[str, Any]:
  workbook = load_workbook(path, data_only=True)
  sheet = workbook["Total"]
  month_columns = {
    (cell_value.year, cell_value.month): col
    for col in range(1, sheet.max_column + 1)
    for cell_value in [sheet.cell(2, col).value]
    if isinstance(cell_value, datetime)
  }

  latest_year, latest_month = max(month_columns)
  previous_year, previous_month = _previous_month(latest_year, latest_month)

  row_map = {
    _clean_label(sheet.cell(row, 1).value): row
    for row in range(1, sheet.max_row + 1)
    if _clean_label(sheet.cell(row, 1).value)
  }

  hp_row = row_map["HP_AIO"]
  asus_row = row_map["ASUS"]
  total_row = row_map["Total"]

  def value(row: int, year: int, month: int) -> int:
    return int(_to_number(sheet.cell(row, month_columns[(year, month)]).value))

  return {
    "asus_current_month": value(asus_row, latest_year, latest_month),
    "asus_previous_month": value(asus_row, previous_year, previous_month),
    "asus_same_month_last_year": value(asus_row, latest_year - 1, latest_month),
    "asus_current_ytd": sum(value(asus_row, latest_year, month) for month in range(1, latest_month + 1)),
    "asus_previous_ytd": sum(value(asus_row, latest_year - 1, month) for month in range(1, latest_month + 1)),
    "hp_current_month": value(hp_row, latest_year, latest_month),
    "hp_same_month_last_year": value(hp_row, latest_year - 1, latest_month),
    "overall_current_month": value(total_row, latest_year, latest_month),
    "overall_same_month_last_year": value(total_row, latest_year - 1, latest_month),
  }


def _load_hp_source(path: Path, latest_year: int, latest_month: int) -> dict[str, Any]:
  workbook = load_workbook(path, data_only=True)
  sheet = workbook["HP-AIO"]
  month_columns = _extract_hp_month_columns(sheet, latest_year, latest_month)
  bps_row = _find_hp_row(sheet, "BPS AIO TTL")
  cps_row = _find_hp_row(sheet, "CPS AIO Total")

  current_year, current_month = latest_year, latest_month
  previous_year, previous_month = _previous_month(current_year, current_month)

  return {
    "bps_current_month": _hp_value(sheet, bps_row, month_columns[(current_year, current_month)]),
    "bps_previous_month": _hp_value(sheet, bps_row, month_columns[(previous_year, previous_month)]),
    "bps_same_month_last_year": _hp_value(sheet, bps_row, month_columns[(current_year - 1, current_month)]),
    "bps_current_ytd": sum(_hp_value(sheet, bps_row, month_columns[(current_year, month)]) for month in range(1, current_month + 1)),
    "bps_previous_ytd": sum(_hp_value(sheet, bps_row, month_columns[(current_year - 1, month)]) for month in range(1, current_month + 1)),
    "cps_current_month": _hp_value(sheet, cps_row, month_columns[(current_year, current_month)]),
    "cps_previous_month": _hp_value(sheet, cps_row, month_columns[(previous_year, previous_month)]),
    "cps_same_month_last_year": _hp_value(sheet, cps_row, month_columns[(current_year - 1, current_month)]),
    "cps_current_ytd": sum(_hp_value(sheet, cps_row, month_columns[(current_year, month)]) for month in range(1, current_month + 1)),
    "cps_previous_ytd": sum(_hp_value(sheet, cps_row, month_columns[(current_year - 1, month)]) for month in range(1, current_month + 1)),
  }


def _extract_hp_month_columns(sheet, latest_year: int, latest_month: int) -> dict[tuple[int, int], int]:
  header_columns = [
    col
    for col in range(1, sheet.max_column + 1)
    if isinstance(sheet.cell(24, col).value, str) and "final" in sheet.cell(24, col).value.lower()
  ]
  if not header_columns:
    raise ValueError("Could not find HP-AIO month header columns.")

  start_year = latest_year - ((len(header_columns) - latest_month) // 12)
  current_year = start_year
  current_month = 1
  month_map: dict[tuple[int, int], int] = {}

  for col in header_columns:
    month_map[(current_year, current_month)] = col
    current_month += 1
    if current_month == 13:
      current_month = 1
      current_year += 1

  if (latest_year, latest_month) not in month_map:
    raise ValueError("Could not align HP-AIO rolling month columns with ASUS latest month.")
  return month_map


def _find_hp_row(sheet, label: str) -> int:
  for row in range(1, sheet.max_row + 1):
    value = sheet.cell(row, 1).value
    if isinstance(value, str) and value.strip().casefold() == label.casefold():
      return row
  raise ValueError(f"Could not find row '{label}' in HP-AIO sheet.")


def _hp_value(sheet, row: int, col: int) -> int:
  return int(_to_number(sheet.cell(row, col).value))


def _group_share(df: pd.DataFrame, field: str, columns: Any) -> dict[str, float]:
  if isinstance(columns, list):
    grouped = (
      df.groupby(field, dropna=False)[columns]
      .sum(numeric_only=True)
      .sum(axis=1)
    )
  else:
    grouped = df.groupby(field, dropna=False)[columns].sum(numeric_only=True)

  total = float(grouped.sum())
  if not total:
    return {}

  result: dict[str, float] = {}
  for label, value in grouped.items():
    normalized_label = _clean_label(label) or "Unknown"
    if normalized_label.lower() == "nan":
      continue
    result[normalized_label] = round(float(value) / total * 100, 2)
  return result


def _with_deltas(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
  output = []
  for item in rows:
    delta = item["current"] - item["previous"]
    output.append(
      {
        **item,
        "delta": delta,
        "ratio": _safe_pct(delta, item["previous"]),
      }
    )
  return output


def _previous_month(year: int, month: int) -> tuple[int, int]:
  if month == 1:
    return year - 1, 12
  return year, month - 1


def _format_month_label(year: int, month: int) -> str:
  return f"{MONTH_SHORT[month - 1]}-{str(year)[-2:]}"


def _to_number(value: Any) -> float:
  if value in (None, "", "-", "--"):
    return 0.0
  if isinstance(value, (int, float)):
    return float(value)
  text = str(value).replace(",", "").replace("\xa0", "").strip()
  if text in ("", "-"):
    return 0.0
  try:
    return float(text)
  except ValueError:
    return 0.0


def _clean_label(value: Any) -> str:
  if value is None:
    return ""
  return str(value).replace("\xa0", " ").strip()


def _safe_pct(delta: int, base: int) -> float | None:
  if not base:
    return None
  return round(delta / base * 100, 2)


def _format_signed(value: int) -> str:
  sign = "+" if value >= 0 else ""
  return f"{sign}{value:,}"


def _format_pct(value: float | None) -> str:
  if value is None:
    return "N/A"
  sign = "+" if value >= 0 else ""
  return f"{sign}{value:.2f}%"
