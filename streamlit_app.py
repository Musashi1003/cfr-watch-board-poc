from __future__ import annotations

import base64
from collections import Counter
from copy import copy
import hmac
from io import BytesIO
import json
import math
import os
import re
from datetime import datetime
from html import escape
from pathlib import Path
from tempfile import NamedTemporaryFile
from urllib import error, request

import altair as alt
import openpyxl
import pandas as pd
import streamlit as st

from cfr_watch_analyzer import (
  FILTER_FIELDS,
  WorkbookUpload,
  analyze_dataset,
  dataset_options,
  parse_workbooks,
)


st.set_page_config(
  page_title="CFR Watch Board",
  layout="wide",
)


FILTER_LABELS = {
  "launch_year": "Launch Year",
  "model": "Model",
  "segment": "Segment",
  "odm_oem": "ODM/OEM",
  "muc_module": "Module",
  "problem_mapping": "Problem",
}

FILTER_ORDER = (
  "launch_year",
  "segment",
  "model",
  "odm_oem",
  "muc_module",
  "problem_mapping",
)
ACT_SCOPE_FILTERS = ("launch_year", "segment", "model", "odm_oem")

BAR_COLOR = "#1a9295"
BAR_LIGHT_COLOR = "#8fd6d5"
LINE_COLOR = "#075f73"
ACCENT_GREEN = "#12805c"
ACCENT_RED = "#d83b35"
TEXT_DARK = "#071316"
GROUP_COMPARE_COLORS = [
  "#075f73",
  "#d83b35",
  "#12805c",
  "#6f4eb2",
  "#b35c00",
  "#0f6fbd",
]
PARSE_CACHE_VERSION = "2026-08-21-act-cache-signature"
ACTIVATION_HISTORY_PATH = Path(__file__).resolve().parent / "data" / "activation_history.csv"
ACTIVATION_HISTORY_COLUMNS = ["source_type", "launch_year", "model", "week", "cumulative_activation", "source"]
ACTIVATION_HISTORY_GITHUB_PATH = "data/activation_history.csv"
ACT_TABLE_PATH = Path(__file__).resolve().parent / "ACT table.xlsx"
ACT_TABLE_GITHUB_PATH = "ACT table.xlsx"
DEFAULT_GITHUB_REPO = "Musashi1003/cfr-watch-board-poc"
DEFAULT_GITHUB_BRANCH = "main"
APP_SESSION_VERSION = "2026-08-21-act-cache-signature"


def week_sort_key(week: str) -> tuple[int, int, str]:
  match = re.search(r"W?(\d{2})(\d{2})", week or "")
  if not match:
    return (0, 0, week or "")
  return (int(match.group(1)), int(match.group(2)), week)


def html_escape(value) -> str:
  return escape(str(value), quote=True)


def read_secret(name: str) -> str:
  value = os.environ.get(name, "")
  if value:
    return value
  try:
    return str(st.secrets.get(name, "") or "")
  except Exception:
    return ""


def password_gate() -> bool:
  expected_username = read_secret("APP_USERNAME").strip()
  expected_password = read_secret("APP_PASSWORD")

  if not expected_password:
    st.error("Access control is not configured. Ask the app owner to set APP_PASSWORD in Streamlit secrets.")
    return False

  if st.session_state.get("authenticated"):
    return True

  st.markdown("### CFR Watch Board Login")
  if expected_username:
    st.info(f"Account: `{expected_username}`. Please ask the dashboard owner for the password.")
  else:
    st.info("Please enter the shared password provided by the dashboard owner.")
  with st.form("login_form"):
    username = st.text_input("Account", value="", disabled=not bool(expected_username))
    password = st.text_input("Password", type="password")
    submitted = st.form_submit_button("Sign in")

  if submitted:
    username_ok = True
    if expected_username:
      username_ok = hmac.compare_digest(username.strip(), expected_username)
    password_ok = hmac.compare_digest(password, expected_password)
    if username_ok and password_ok:
      st.session_state["authenticated"] = True
      st.rerun()
    st.error("Account or password is incorrect.")

  return False


def reset_session_if_app_version_changed():
  if st.session_state.get("app_session_version") == APP_SESSION_VERSION:
    return
  st.session_state.clear()
  st.session_state["app_session_version"] = APP_SESSION_VERSION


def apply_page_style():
  st.markdown(
    """
    <style>
      .block-container {
        padding-top: 1.8rem;
        padding-bottom: 2.5rem;
        max-width: 1480px;
      }
      .stApp {
        background: #f6f9fb;
      }
      [data-testid="stSidebar"] {
        background: #f8fbfc;
        border-right: 1px solid #dce8ea;
      }
      [data-testid="stSidebar"] h1,
      [data-testid="stSidebar"] h2,
      [data-testid="stSidebar"] h3,
      [data-testid="stSidebar"] p,
      [data-testid="stSidebar"] label {
        color: #122426;
      }
      [data-testid="stSidebar"] [data-baseweb="select"] > div {
        background: #ffffff;
        border-color: #cddbdd;
      }
      [data-testid="stSidebar"] [data-baseweb="select"] span,
      [data-testid="stSidebar"] [data-baseweb="select"] input {
        color: #122426 !important;
      }
      [data-testid="stSidebar"] [data-baseweb="tag"] {
        background-color: #dff3f3 !important;
        border-color: #b6dfdf !important;
      }
      [data-testid="stSidebar"] [data-baseweb="tag"] span {
        color: #075f73 !important;
        font-weight: 650;
      }
      .cfr-title {
        color: #122426;
        font-size: 2.1rem;
        font-weight: 780;
        margin-bottom: 0.15rem;
      }
      .cfr-subtitle {
        color: #5f6f72;
        font-size: 0.98rem;
        margin-bottom: 1.1rem;
      }
      .dashboard-head {
        display: flex;
        align-items: flex-end;
        justify-content: space-between;
        gap: 1rem;
        margin: 1rem 0 0.8rem 0;
      }
      .dashboard-head h2 {
        font-size: 1.45rem;
        margin: 0 0 0.2rem 0;
      }
      .dashboard-head p {
        color: #62767b;
        margin: 0;
      }
      .mode-head {
        margin-top: 1.25rem;
        margin-bottom: 0.55rem;
      }
      .mode-note {
        color: #557179;
        font-size: 0.85rem;
        white-space: nowrap;
      }
      .stButton > button {
        border-radius: 8px;
        min-height: 2.55rem;
        font-weight: 700;
      }
      .stButton > button[kind="primary"] {
        background: #075f73;
        border-color: #075f73;
        color: #ffffff;
      }
      .stButton > button[kind="primary"]:hover {
        background: #064f60;
        border-color: #064f60;
        color: #ffffff;
      }
      .stButton > button[kind="secondary"] {
        background: #ffffff;
        border-color: #cddbdd;
        color: #122426;
      }
      .stButton > button[kind="secondary"]:hover {
        border-color: #138a8e;
        color: #075f73;
      }
      .refresh-note {
        color: #557179;
        font-size: 0.85rem;
        white-space: nowrap;
      }
      .metric-card {
        background: linear-gradient(135deg, #ffffff 0%, #f9fffd 100%);
        border: 1px solid #e5e0d8;
        border-radius: 8px;
        padding: 1rem 1.1rem;
        min-height: 150px;
        box-shadow: 0 10px 24px rgba(16, 35, 42, 0.06);
        border-top: 4px solid var(--accent, #138a8e);
      }
      .metric-label {
        color: #4d6568;
        font-size: 0.82rem;
        margin-bottom: 0.45rem;
      }
      .metric-value {
        color: #071316;
        font-size: 1.75rem;
        font-weight: 800;
        line-height: 1.1;
      }
      .metric-note {
        color: #607478;
        font-size: 0.82rem;
        margin-top: 0.65rem;
      }
      .metric-delta {
        display: inline-block;
        color: var(--accent, #138a8e);
        font-size: 0.78rem;
        font-weight: 700;
        margin-top: 0.75rem;
      }
      .sparkline {
        width: 100%;
        height: 32px;
        margin-top: 0.75rem;
      }
      h2, h3 {
        color: #0c1c1f;
      }
      div[data-testid="stDataFrame"] {
        border: 1px solid #e5e0d8;
        border-radius: 8px;
      }
      .action-card {
        background: linear-gradient(135deg, #ffffff 0%, #f9fffd 100%);
        border: 1px solid #e5e0d8;
        border-top: 4px solid #1a9295;
        border-radius: 8px;
        padding: 1rem 1.1rem;
        min-height: 152px;
        box-shadow: 0 10px 24px rgba(16, 35, 42, 0.06);
      }
      .action-label {
        color: #4d6568;
        font-size: 0.78rem;
        font-weight: 750;
        letter-spacing: 0.04em;
        margin-bottom: 0.45rem;
      }
      .action-title {
        color: #071316;
        font-size: 1rem;
        font-weight: 800;
        line-height: 1.25;
        overflow-wrap: anywhere;
        min-height: 2.5rem;
      }
      .action-number {
        display: flex;
        align-items: baseline;
        gap: 0.55rem;
        margin-top: 0.85rem;
      }
      .action-number strong {
        color: #071316;
        font-size: 1.75rem;
        line-height: 1;
      }
      .action-number span {
        color: #607478;
        font-size: 0.86rem;
      }
      .action-note {
        color: #607478;
        font-size: 0.8rem;
        margin-top: 0.65rem;
        overflow-wrap: anywhere;
      }
      .change-log {
        background: #ffffff;
        border: 1px solid #e5e0d8;
        border-radius: 8px;
        padding: 1rem 1.15rem;
        box-shadow: 0 10px 24px rgba(16, 35, 42, 0.06);
      }
      .change-log h3 {
        margin: 0 0 0.7rem 0;
        font-size: 1.08rem;
      }
      .change-log ol {
        margin: 0;
        padding-left: 1.2rem;
      }
      .change-log li {
        color: #5f6f72;
        font-size: 0.88rem;
        line-height: 1.7;
      }
      .change-log strong {
        color: #122426;
      }
    </style>
    """,
    unsafe_allow_html=True,
  )


def pct(value: float | None) -> str:
  if value is None:
    return "N/A"
  return f"{value * 100:.2f}%"


def whole(value: float | int | None) -> str:
  if value is None:
    return "N/A"
  return f"{value:,.0f}"


def nice_axis_values(max_value: float | int | None) -> list[int]:
  if not max_value or max_value <= 0:
    return [0]

  if max_value <= 2:
    upper = int(math.ceil(max_value))
    return [0, upper] if upper == 1 else [0, 1, upper]

  target = max_value * 1.05
  exponent = math.floor(math.log10(target))
  base = 10 ** exponent
  fraction = target / base
  if fraction <= 1:
    nice_fraction = 1
  elif fraction <= 2:
    nice_fraction = 2
  elif fraction <= 5:
    nice_fraction = 5
  else:
    nice_fraction = 10

  upper = int(nice_fraction * base)
  middle = int(upper / 2)
  return [0, middle, upper]


def sparkline_svg(values: list[int], color: str) -> str:
  if len(values) < 2:
    return ""

  width = 220
  height = 34
  padding = 3
  min_value = min(values)
  max_value = max(values)
  span = max(max_value - min_value, 1)
  points = []
  for index, value in enumerate(values):
    x = padding + index * ((width - padding * 2) / (len(values) - 1))
    y = height - padding - ((value - min_value) / span) * (height - padding * 2)
    points.append(f"{x:.1f},{y:.1f}")
  return (
    f'<svg class="sparkline" viewBox="0 0 {width} {height}" preserveAspectRatio="none">'
    f'<polyline points="{" ".join(points)}" fill="none" stroke="{color}" stroke-width="3" '
    'stroke-linecap="round" stroke-linejoin="round" />'
    f'<circle cx="{points[-1].split(",")[0]}" cy="{points[-1].split(",")[1]}" r="3.2" fill="{color}" />'
    '</svg>'
  )


def dashboard_header(result: dict):
  latest_week = result["kpis"]["latest_week"]
  st.markdown(
    f"""
    <div class="dashboard-head">
      <div>
        <h2>Overview Dashboard</h2>
      </div>
      <div class="refresh-note">Updated: {datetime.now().strftime("%Y-%m-%d %H:%M")} | Latest week: {latest_week}</div>
    </div>
    """,
    unsafe_allow_html=True,
  )


def render_mode_selector() -> str:
  current_mode = st.session_state.get("view_mode", "Dashboard")
  current_label = "Overview Dashboard" if current_mode == "Dashboard" else "Group CFR Compare"
  st.markdown(
    f"""
    <div class="dashboard-head mode-head">
      <div>
        <h2>Analysis Mode</h2>
      </div>
      <div class="mode-note">Current: {current_label}</div>
    </div>
    """,
    unsafe_allow_html=True,
  )

  dashboard_col, compare_col, _ = st.columns([1.15, 1.2, 5.2])
  with dashboard_col:
    if st.button(
      "Overview Dashboard",
      type="primary" if current_mode == "Dashboard" else "secondary",
      width="stretch",
    ):
      st.session_state["view_mode"] = "Dashboard"
      st.rerun()
  with compare_col:
    if st.button(
      "Group CFR Compare",
      type="primary" if current_mode == "Group Compare" else "secondary",
      width="stretch",
    ):
      st.session_state["view_mode"] = "Group Compare"
      st.rerun()

  return st.session_state.get("view_mode", "Dashboard")


def write_uploads_to_temp_files(uploaded_files) -> tuple[list[WorkbookUpload], list[Path]]:
  workbooks: list[WorkbookUpload] = []
  temp_paths: list[Path] = []
  for filename, file_bytes in uploaded_files:
    suffix = Path(filename).suffix
    with NamedTemporaryFile(delete=False, suffix=suffix) as temp_file:
      temp_file.write(file_bytes)
      temp_path = Path(temp_file.name)
    temp_paths.append(temp_path)
    workbooks.append(WorkbookUpload(path=temp_path, filename=filename))
  return workbooks, temp_paths


def build_upload_payloads(uploaded_files) -> tuple[tuple[str, bytes], ...]:
  return tuple((uploaded_file.name, uploaded_file.getvalue()) for uploaded_file in uploaded_files)


def uploaded_size_mb(upload_payloads: tuple[tuple[str, bytes], ...]) -> float:
  return sum(len(file_bytes) for _, file_bytes in upload_payloads) / (1024 * 1024)


def upload_summary_text(upload_payloads: tuple[tuple[str, bytes], ...]) -> str:
  total_mb = uploaded_size_mb(upload_payloads)
  file_count = len(upload_payloads)
  return f"Received {file_count} workbook{'s' if file_count != 1 else ''}, total {total_mb:.1f} MB."


@st.cache_data(show_spinner=False, ttl=3600, max_entries=8)
def parse_uploaded_payloads(upload_payloads: tuple[tuple[str, bytes], ...], parser_version: str) -> dict:
  _ = parser_version
  workbooks, temp_paths = write_uploads_to_temp_files(upload_payloads)
  try:
    return parse_workbooks(workbooks)
  finally:
    for temp_path in temp_paths:
      try:
        temp_path.unlink(missing_ok=True)
      except PermissionError:
        pass


def latest_week_from_records(records: list[dict]) -> str:
  weeks = sorted(
    {str(record.get("Week", "")).strip() for record in records if str(record.get("Week", "")).strip()},
    key=week_sort_key,
  )
  return weeks[-1] if weeks else ""


def normalize_source_type(value) -> str:
  upper_value = str(value or "").upper().replace("_", "-").replace(" ", "-")
  if "GAMING" in upper_value:
    return "Gaming NB"
  if "PC" in upper_value:
    return "PC NB"
  return str(value or "").strip() or "Uploaded"


def format_activation_value(value: float) -> str:
  if abs(value - round(value)) < 0.0001:
    return str(int(round(value)))
  return f"{value:.2f}".rstrip("0").rstrip(".")


def act_table_source_type(value) -> str:
  source_type = normalize_source_type(value)
  if source_type == "Gaming NB":
    return "GAMING-NB"
  if source_type == "PC NB":
    return "PC-NB"
  return str(value or "").strip() or "Uploaded"


def clean_launch_year(value) -> str:
  match = re.search(r"(20\d{2})", str(value or ""))
  return match.group(1) if match else ""


def launch_year_from_record(record: dict) -> str:
  for key in ("\u958b\u8ce3\u5e74\u5ea6", "LAUNCH_YEAR", "OPEN_YEAR", "SALE_YEAR"):
    launch_year = clean_launch_year(record.get(key, ""))
    if launch_year:
      return launch_year
  return ""


def act_model_from_record(record: dict) -> str:
  launch_year = launch_year_from_record(record)
  if launch_year == "2025":
    model_group = str(record.get("MODEL_GROUP", "")).strip()
    if model_group:
      return model_group
  return str(record.get("ORG_MODEL(PRODUCT_DESC)", "")).strip()


def normalized_identifier(value: str) -> str:
  return re.sub(r"[^A-Z0-9]", "", str(value or "").upper())


def summary_key_for_model(model: str, summary_by_model: dict[str, dict]) -> str | None:
  if model in summary_by_model:
    return model

  model_key = normalized_identifier(model)
  if not model_key:
    return None

  prefix_matches = [
    summary_model
    for summary_model in summary_by_model
    if len(normalized_identifier(summary_model)) >= 4
    and model_key.startswith(normalized_identifier(summary_model))
  ]
  if not prefix_matches:
    return None
  return max(prefix_matches, key=lambda summary_model: len(normalized_identifier(summary_model)))


def configured_act_table_path() -> Path:
  configured_path = read_secret("ACT_TABLE_PATH").strip()
  if configured_path:
    return Path(configured_path)

  if ACT_TABLE_PATH.exists():
    return ACT_TABLE_PATH

  local_fallback = Path(r"C:\=Codex study==\20260604-Sites in Codex\ACT table.xlsx")
  if local_fallback.exists():
    return local_fallback

  return ACT_TABLE_PATH


def file_signature(path: Path) -> tuple[str, int, int] | None:
  try:
    stat = path.stat()
  except OSError:
    return None
  return (str(path), int(stat.st_mtime), int(stat.st_size))


def act_store_cache_signature() -> tuple:
  return (
    file_signature(ACTIVATION_HISTORY_PATH),
    file_signature(configured_act_table_path()),
    APP_SESSION_VERSION,
  )


def numeric_activation(value) -> float | None:
  if value is None or value == "":
    return None
  if isinstance(value, (int, float)):
    return float(value)
  text = str(value).strip().replace(",", "")
  if not text:
    return None
  try:
    return float(text)
  except ValueError:
    return None


def load_act_table_store() -> dict[tuple[str, str, str, str], float]:
  path = configured_act_table_path()
  if not path.exists():
    return {}

  store: dict[tuple[str, str, str, str], float] = {}
  workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
  try:
    for worksheet in workbook.worksheets:
      header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
      if not header_row:
        continue
      headers = [str(value or "").strip() for value in header_row]
      normalized_headers = {normalized_identifier(header): index for index, header in enumerate(headers)}
      source_index = normalized_headers.get("GAMINGPC")
      model_index = (
        normalized_headers.get("ORGMODELPRODUCTDESC")
        or normalized_headers.get("MODELGROUP")
        or normalized_headers.get("MODEL")
      )
      if source_index is None or model_index is None:
        continue

      week_indices = [
        (index, header)
        for index, header in enumerate(headers)
        if re.fullmatch(r"W\d{4}", header, flags=re.IGNORECASE)
      ]
      if not week_indices:
        continue

      sheet_launch_year = clean_launch_year(worksheet.title)
      year_index = 0 if headers and ("年度" in headers[0] or "YEAR" in headers[0].upper()) else None
      for row in worksheet.iter_rows(min_row=2, values_only=True):
        source_type = normalize_source_type(row[source_index] if source_index < len(row) else "")
        launch_year = sheet_launch_year
        if year_index is not None and year_index < len(row):
          launch_year = clean_launch_year(row[year_index]) or launch_year
        model = str(row[model_index] if model_index < len(row) else "").strip()
        if not source_type or not launch_year or not model:
          continue
        for week_index, week in week_indices:
          activation = numeric_activation(row[week_index] if week_index < len(row) else None)
          if activation is None:
            continue
          store[(source_type, launch_year, model, week.upper())] = activation
  finally:
    workbook.close()
  return store


@st.cache_resource
def activation_history_store_cached(cache_signature: tuple) -> dict[tuple[str, str, str, str], float]:
  _ = cache_signature
  store: dict[tuple[str, str, str, str], float] = {}
  if ACTIVATION_HISTORY_PATH.exists():
    history = pd.read_csv(ACTIVATION_HISTORY_PATH)
    for row in history.to_dict("records"):
      source_type = normalize_source_type(row.get("source_type", ""))
      launch_year = clean_launch_year(row.get("launch_year", ""))
      model = str(row.get("model", "")).strip()
      week = str(row.get("week", "")).strip()
      if not model or not week:
        continue
      try:
        activation = float(row.get("cumulative_activation", 0) or 0)
      except (TypeError, ValueError):
        activation = 0.0
      store[(source_type, launch_year, model, week)] = activation

  store.update(load_act_table_store())
  return store


def activation_history_store() -> dict[tuple[str, str, str, str], float]:
  return activation_history_store_cached(act_store_cache_signature())


def activation_updates_from_parsed(parsed: dict) -> list[dict]:
  records = parsed.get("records", [])
  week = latest_week_from_records(records)
  if not week:
    return []

  summary_by_model = parsed.get("summary_by_model", {})
  scope_by_act_model: dict[tuple[str, str, str], set[str]] = {}
  for record in records:
    source_type = normalize_source_type(record.get("source_type", ""))
    launch_year = launch_year_from_record(record)
    act_model = act_model_from_record(record)
    raw_model = str(record.get("ORG_MODEL(PRODUCT_DESC)", "")).strip()
    if not source_type or not launch_year or not act_model or not raw_model:
      continue
    scope_by_act_model.setdefault((source_type, launch_year, act_model), set()).add(raw_model)

  updates = []
  for (source_type, launch_year, act_model), raw_models in sorted(scope_by_act_model.items()):
    activation = 0.0
    matched_models = 0
    for raw_model in raw_models:
      summary_key = summary_key_for_model(raw_model, summary_by_model)
      if not summary_key:
        continue
      derived_act = summary_by_model[summary_key].get("derived_act")
      if derived_act is None:
        continue
      activation += float(derived_act)
      matched_models += 1

    if not matched_models:
      continue

    updates.append(
      {
        "source_type": source_type,
        "launch_year": launch_year,
        "model": act_model,
        "week": week,
        "cumulative_activation": activation,
        "source": "upload",
      }
    )
  return updates


def merge_activation_history(updates: list[dict]) -> tuple[pd.DataFrame, int, int]:
  if ACTIVATION_HISTORY_PATH.exists():
    history = pd.read_csv(ACTIVATION_HISTORY_PATH, dtype=str).fillna("")
  else:
    history = pd.DataFrame(columns=ACTIVATION_HISTORY_COLUMNS)

  for column in ACTIVATION_HISTORY_COLUMNS:
    if column not in history.columns:
      history[column] = ""
  history = history[ACTIVATION_HISTORY_COLUMNS].copy()

  existing_index = {
    (
      normalize_source_type(row["source_type"]),
      clean_launch_year(row.get("launch_year", "")),
      str(row["model"]).strip(),
      str(row["week"]).strip(),
    ): index
    for index, row in history.iterrows()
  }

  for (source_type, launch_year, model, week), activation in activation_history_store().items():
    key = (source_type, launch_year, model, week)
    activation_text = format_activation_value(activation)
    if key in existing_index:
      history.at[existing_index[key], "cumulative_activation"] = activation_text
      continue
    history.loc[len(history)] = {
      "source_type": source_type,
      "launch_year": launch_year,
      "model": model,
      "week": week,
      "cumulative_activation": activation_text,
      "source": "upload",
    }
    existing_index[key] = len(history) - 1

  added_count = 0
  changed_count = 0
  act_table_store = load_act_table_store()
  for update in updates:
    key = (update["source_type"], update["launch_year"], update["model"], update["week"])
    if key in act_table_store:
      activation_text = format_activation_value(act_table_store[key])
      if key in existing_index:
        row_index = existing_index[key]
        history.at[row_index, "cumulative_activation"] = activation_text
        history.at[row_index, "source"] = "act_table"
      else:
        history.loc[len(history)] = {
          "source_type": update["source_type"],
          "launch_year": update["launch_year"],
          "model": update["model"],
          "week": update["week"],
          "cumulative_activation": activation_text,
          "source": "act_table",
        }
        existing_index[key] = len(history) - 1
      continue

    activation_text = format_activation_value(update["cumulative_activation"])
    if key in existing_index:
      row_index = existing_index[key]
      old_value = str(history.at[row_index, "cumulative_activation"]).strip()
      if old_value != activation_text:
        history.at[row_index, "cumulative_activation"] = activation_text
        history.at[row_index, "source"] = "upload"
        changed_count += 1
      continue

    history.loc[len(history)] = {
      "source_type": update["source_type"],
      "launch_year": update["launch_year"],
      "model": update["model"],
      "week": update["week"],
      "cumulative_activation": activation_text,
      "source": "upload",
    }
    existing_index[key] = len(history) - 1
    added_count += 1

  history["_source_order"] = history["source_type"].map({"PC NB": 0, "Gaming NB": 1}).fillna(9)
  history["_week_order"] = history["week"].map(week_sort_key)
  history = history.sort_values(["_source_order", "launch_year", "model", "_week_order"]).drop(columns=["_source_order", "_week_order"])
  return history[ACTIVATION_HISTORY_COLUMNS], added_count, changed_count


def activation_history_csv_text(history: pd.DataFrame) -> str:
  return history.to_csv(index=False, lineterminator="\n")


def github_request(url: str, method: str, token: str, payload: dict | None = None) -> dict:
  body = json.dumps(payload).encode("utf-8") if payload is not None else None
  api_request = request.Request(url, data=body, method=method)
  api_request.add_header("Accept", "application/vnd.github+json")
  api_request.add_header("Authorization", f"Bearer {token}")
  api_request.add_header("X-GitHub-Api-Version", "2022-11-28")
  if payload is not None:
    api_request.add_header("Content-Type", "application/json")

  with request.urlopen(api_request, timeout=20) as response:
    return json.loads(response.read().decode("utf-8"))


def write_activation_history_to_github(csv_text: str, changed_count: int, added_count: int) -> tuple[bool, str]:
  token = read_secret("ACT_HISTORY_GITHUB_TOKEN") or read_secret("GITHUB_TOKEN")
  if not token:
    return False, "GitHub token is not configured."

  repo = read_secret("ACT_HISTORY_GITHUB_REPO") or DEFAULT_GITHUB_REPO
  branch = read_secret("ACT_HISTORY_GITHUB_BRANCH") or DEFAULT_GITHUB_BRANCH
  path = read_secret("ACT_HISTORY_GITHUB_PATH") or ACTIVATION_HISTORY_GITHUB_PATH
  contents_url = f"https://api.github.com/repos/{repo}/contents/{path}"

  try:
    current_file = github_request(f"{contents_url}?ref={branch}", "GET", token)
    commit_message = f"Update activation history ({added_count} added, {changed_count} changed)"
    github_request(
      contents_url,
      "PUT",
      token,
      {
        "message": commit_message,
        "content": base64.b64encode(csv_text.encode("utf-8")).decode("ascii"),
        "branch": branch,
        "sha": current_file["sha"],
      },
    )
  except error.HTTPError as exc:
    detail = exc.read().decode("utf-8", errors="replace")
    return False, f"GitHub write failed: HTTP {exc.code} {detail[:220]}"
  except Exception as exc:
    return False, f"GitHub write failed: {exc}"

  return True, f"Saved to GitHub {path} on {branch}."

def write_bytes_to_github(path: str, content: bytes, commit_message: str) -> tuple[bool, str]:
  token = read_secret("ACT_HISTORY_GITHUB_TOKEN") or read_secret("GITHUB_TOKEN")
  if not token:
    return False, "GitHub token is not configured."

  repo = read_secret("ACT_HISTORY_GITHUB_REPO") or DEFAULT_GITHUB_REPO
  branch = read_secret("ACT_HISTORY_GITHUB_BRANCH") or DEFAULT_GITHUB_BRANCH
  contents_url = f"https://api.github.com/repos/{repo}/contents/{path}"

  payload = {
    "message": commit_message,
    "content": base64.b64encode(content).decode("ascii"),
    "branch": branch,
  }
  try:
    current_file = github_request(f"{contents_url}?ref={branch}", "GET", token)
    payload["sha"] = current_file["sha"]
  except error.HTTPError as exc:
    if exc.code != 404:
      detail = exc.read().decode("utf-8", errors="replace")
      return False, f"GitHub write failed: HTTP {exc.code} {detail[:220]}"
  except Exception as exc:
    return False, f"GitHub write failed: {exc}"

  try:
    github_request(contents_url, "PUT", token, payload)
  except error.HTTPError as exc:
    detail = exc.read().decode("utf-8", errors="replace")
    return False, f"GitHub write failed: HTTP {exc.code} {detail[:220]}"
  except Exception as exc:
    return False, f"GitHub write failed: {exc}"

  return True, f"Saved to GitHub {path} on {branch}."


def act_github_write_enabled() -> bool:
  return bool(read_secret("ACT_HISTORY_GITHUB_TOKEN") or read_secret("GITHUB_TOKEN"))


def act_update_weeks(updates: list[dict]) -> list[str]:
  return sorted(
    {str(update.get("week", "")).strip() for update in updates if str(update.get("week", "")).strip()},
    key=week_sort_key,
  )


def act_update_week_label(updates: list[dict]) -> str:
  weeks = act_update_weeks(updates)
  if not weeks:
    return "N/A"
  if len(weeks) == 1:
    return weeks[0]
  return f"{weeks[0]}-{weeks[-1]}"


def worksheet_header_map(worksheet) -> dict[str, int]:
  header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
  if not header_row:
    return {}
  return {
    normalized_identifier(value): index + 1
    for index, value in enumerate(header_row)
    if str(value or "").strip()
  }


def ensure_act_sheet(workbook, launch_year: str):
  sheet_name = f"{launch_year} ACT"
  if sheet_name in workbook.sheetnames:
    return workbook[sheet_name]

  worksheet = workbook.create_sheet(sheet_name)
  model_header = "MODEL_GROUP" if launch_year == "2025" else "ORG_MODEL(PRODUCT_DESC)"
  worksheet.append(["\u958b\u8ce3\u5e74\u5ea6", "GAMING/PC", model_header])
  return worksheet


def ensure_week_column(worksheet, week: str) -> int:
  headers = worksheet_header_map(worksheet)
  week_key = normalized_identifier(week)
  if week_key in headers:
    return headers[week_key]

  new_column = worksheet.max_column + 1
  worksheet.cell(row=1, column=new_column).value = week
  if new_column > 1:
    source_cell = worksheet.cell(row=1, column=new_column - 1)
    target_cell = worksheet.cell(row=1, column=new_column)
    if source_cell.has_style:
      target_cell._style = copy(source_cell._style)
    target_cell.number_format = source_cell.number_format
    target_cell.alignment = copy(source_cell.alignment)
  return new_column


def act_table_row_key(worksheet, row_index: int, source_column: int, model_column: int) -> tuple[str, str]:
  source_type = normalize_source_type(worksheet.cell(row=row_index, column=source_column).value)
  model = str(worksheet.cell(row=row_index, column=model_column).value or "").strip()
  return source_type, model


def find_or_create_act_row(worksheet, update: dict, source_column: int, model_column: int) -> int:
  target_key = (update["source_type"], update["model"])
  for row_index in range(2, worksheet.max_row + 1):
    if act_table_row_key(worksheet, row_index, source_column, model_column) == target_key:
      return row_index

  row_index = worksheet.max_row + 1
  worksheet.cell(row=row_index, column=1).value = int(update["launch_year"])
  worksheet.cell(row=row_index, column=source_column).value = act_table_source_type(update["source_type"])
  worksheet.cell(row=row_index, column=model_column).value = update["model"]
  if row_index > 2:
    for column in range(1, worksheet.max_column + 1):
      source_cell = worksheet.cell(row=row_index - 1, column=column)
      target_cell = worksheet.cell(row=row_index, column=column)
      if source_cell.has_style:
        target_cell._style = copy(source_cell._style)
      target_cell.number_format = source_cell.number_format
      target_cell.alignment = copy(source_cell.alignment)
  return row_index


def update_act_table_workbook(updates: list[dict]) -> dict:
  if not updates:
    return {
      "status": "skipped",
      "message": "No ACT values were found to record.",
      "persistence_ok": True,
      "requires_manual_save": False,
    }

  path = configured_act_table_path()
  if path.exists():
    workbook = openpyxl.load_workbook(path)
  else:
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)

  added_count = 0
  kept_count = 0
  try:
    for update in updates:
      launch_year = str(update.get("launch_year", "")).strip()
      if not launch_year:
        continue
      worksheet = ensure_act_sheet(workbook, launch_year)
      headers = worksheet_header_map(worksheet)
      source_column = headers.get("GAMINGPC") or 2
      model_column = (
        headers.get("ORGMODELPRODUCTDESC")
        or headers.get("MODELGROUP")
        or headers.get("MODEL")
        or 3
      )
      week_column = ensure_week_column(worksheet, update["week"])
      row_index = find_or_create_act_row(worksheet, update, source_column, model_column)
      target_cell = worksheet.cell(row=row_index, column=week_column)
      existing_value = numeric_activation(target_cell.value)
      if existing_value is not None:
        kept_count += 1
        continue
      target_cell.value = int(round(update["cumulative_activation"]))
      target_cell.number_format = "#,##0"
      added_count += 1

    output = BytesIO()
    workbook.save(output)
  finally:
    workbook.close()

  if added_count == 0:
    return {
      "status": "unchanged",
      "message": f"ACT table already has the latest uploaded week values. Kept {kept_count} existing values.",
      "added_count": added_count,
      "kept_count": kept_count,
      "generated_count": len(updates),
      "week_label": act_update_week_label(updates),
      "github_write_enabled": act_github_write_enabled(),
      "persistence_ok": True,
      "requires_manual_save": False,
    }

  xlsx_bytes = output.getvalue()
  saved, message = write_bytes_to_github(
    read_secret("ACT_TABLE_GITHUB_PATH") or ACT_TABLE_GITHUB_PATH,
    xlsx_bytes,
    f"Update ACT table ({added_count} values)",
  )
  return {
    "status": "saved" if saved else "download",
    "message": message,
    "added_count": added_count,
    "kept_count": kept_count,
    "generated_count": len(updates),
    "week_label": act_update_week_label(updates),
    "github_write_enabled": act_github_write_enabled(),
    "persistence_ok": saved,
    "requires_manual_save": not saved,
    "xlsx_bytes": xlsx_bytes,
  }


def remember_activation_snapshot(parsed: dict) -> dict:
  updates = activation_updates_from_parsed(parsed)
  if not updates:
    return {"status": "skipped", "message": "No ACT values were found to record."}

  merged_history, added_count, changed_count = merge_activation_history(updates)
  store = activation_history_store()
  act_table_store = load_act_table_store()
  for update in updates:
    key = (update["source_type"], update["launch_year"], update["model"], update["week"])
    store[key] = act_table_store.get(key, update["cumulative_activation"])

  csv_text = activation_history_csv_text(merged_history)
  if added_count == 0 and changed_count == 0:
    return {"status": "unchanged", "message": "Activation history is already up to date."}

  saved, message = write_activation_history_to_github(csv_text, changed_count, added_count)
  return {
    "status": "saved" if saved else "download",
    "message": message,
    "added_count": added_count,
    "changed_count": changed_count,
    "csv_text": csv_text,
  }


def remember_act_table_snapshot(parsed: dict) -> dict:
  return update_act_table_workbook(activation_updates_from_parsed(parsed))


def selected_filters(records: list[dict]) -> dict[str, list[str]]:
  current_selections = {
    key: [
      str(value).strip()
      for value in st.session_state.get(f"filter_{key}", [])
      if str(value).strip()
    ]
    for key in FILTER_FIELDS
  }
  options_by_filter = dataset_options(records, current_selections)
  selections: dict[str, list[str]] = {}
  with st.sidebar:
    st.header("Filters")
    st.caption("Leave a filter empty to include all values.")
    for key in FILTER_ORDER:
      options = options_by_filter.get(key, [])
      with st.expander(FILTER_LABELS[key], expanded=False):
        selections[key] = st.multiselect(
          "Select values",
          options,
          default=current_selections.get(key, []),
          key=f"filter_{key}",
          label_visibility="collapsed",
        )
  return selections


def filter_controls(records: list[dict], key_prefix: str) -> dict[str, list[str]]:
  current_selections = {
    key: [
      str(value).strip()
      for value in st.session_state.get(f"{key_prefix}_{key}", [])
      if str(value).strip()
    ]
    for key in FILTER_FIELDS
  }
  options_by_filter = dataset_options(records, current_selections)
  selections: dict[str, list[str]] = {}
  for key in FILTER_ORDER:
    options = options_by_filter.get(key, [])
    selections[key] = st.multiselect(
      FILTER_LABELS[key],
      options,
      default=current_selections.get(key, []),
      key=f"{key_prefix}_{key}",
    )
  return selections


def clean_filter_snapshot(filters: dict[str, list[str]]) -> dict[str, list[str]]:
  return {
    key: [str(value).strip() for value in filters.get(key, []) if str(value).strip()]
    for key in FILTER_FIELDS
  }


def filter_snapshot_summary(filters: dict[str, list[str]]) -> str:
  parts = []
  for key in FILTER_ORDER:
    values = filters.get(key, [])
    if not values:
      continue
    label = FILTER_LABELS[key]
    preview = ", ".join(values[:2])
    if len(values) > 2:
      preview = f"{preview}, +{len(values) - 2} more"
    parts.append(f"{label}: {preview}")
  return " | ".join(parts) if parts else "All records"

def filter_snapshot_details(filters: dict[str, list[str]]) -> pd.DataFrame:
  rows = []
  for key in FILTER_ORDER:
    values = filters.get(key, [])
    rows.append(
      {
        "Filter": FILTER_LABELS[key],
        "Selected Values": ", ".join(values) if values else "All",
      }
    )
  return pd.DataFrame(rows)


def normalized_filter_signature(filters: dict[str, list[str]]) -> tuple[tuple[str, tuple[str, ...]], ...]:
  snapshot = clean_filter_snapshot(filters)
  return tuple(
    (key, tuple(sorted(snapshot.get(key, []))))
    for key in FILTER_ORDER
  )


def matching_group_label(filters: dict[str, list[str]], groups: list[dict]) -> str | None:
  signature = normalized_filter_signature(filters)
  for group in groups:
    if normalized_filter_signature(group.get("filters", {})) == signature:
      return str(group.get("label", "")).strip() or "another group"
  return None


def render_activation_history_status(result: dict):
  status = result.get("status")
  message = result.get("message", "")
  added_count = result.get("added_count", 0)
  changed_count = result.get("changed_count", 0)

  if status == "saved":
    st.success(f"ACT CSV history saved: added {added_count}, changed {changed_count}. {message}")
    return
  if status == "unchanged":
    st.info("ACT CSV history is already up to date.")
    return
  if status == "skipped":
    st.info(message or "No ACT values were found for CSV history.")
    return
  if status == "download":
    st.warning(
      f"ACT CSV history was updated in this session but was not saved to GitHub. {message} "
      "Download the CSV if you still need the legacy history file."
    )
    st.download_button(
      "Download updated activation_history.csv",
      data=result.get("csv_text", ""),
      file_name="activation_history.csv",
      mime="text/csv",
    )


def render_act_table_status(result: dict):
  status = result.get("status")
  message = result.get("message", "")
  added_count = result.get("added_count", 0)
  kept_count = result.get("kept_count", 0)
  generated_count = result.get("generated_count", 0)
  week_label = result.get("week_label", "N/A")

  if status == "saved":
    st.success(
      f"ACT Persistence Guard passed: {week_label} saved to ACT table "
      f"({added_count} new values, {kept_count} kept). {message}"
    )
    return
  if status == "unchanged":
    st.success(
      message
      or f"ACT Persistence Guard passed: {week_label} already exists in ACT table."
    )
    return
  if status == "skipped":
    st.info(message or "No ACT values were found for ACT table update.")
    return
  if status == "download":
    st.error(
      f"ACT Persistence Guard failed: {week_label} generated {added_count} new ACT values "
      f"from {generated_count} model updates, but they were not permanently saved. {message}"
    )
    st.warning(
      "Download the updated ACT table below and replace `ACT table.xlsx`, or ask the app owner to set "
      "`ACT_HISTORY_GITHUB_TOKEN` (or `GITHUB_TOKEN`) in Streamlit Secrets so future uploads can save automatically."
    )
    st.download_button(
      "Download updated ACT table.xlsx",
      data=result.get("xlsx_bytes", b""),
      file_name="ACT table.xlsx",
      mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def render_act_persistence_preflight():
  if act_github_write_enabled():
    st.sidebar.success("ACT persistence: GitHub write enabled.")
    return

  st.sidebar.warning(
    "ACT persistence: GitHub write token is not configured. "
    "New ACT weeks can be calculated, but they will not be saved permanently unless the updated ACT table is downloaded and committed."
  )


def activation_trend_values(records: list[dict], filters: dict[str, list[str]], max_points: int = 12) -> list[int]:
  store = activation_history_store()
  if not store:
    return []

  model_scope = model_scope_for_filters(records, filters)
  if not model_scope:
    return []

  latest_uploaded_week = latest_week_from_records(records)
  available_weeks = sorted(
    {
      week
      for source_type, launch_year, model in model_scope
      for stored_source, stored_year, stored_model, week in store
      if stored_model == model and (stored_source == source_type or source_type == "Uploaded")
      if stored_year in ("", launch_year)
      if not latest_uploaded_week or week_sort_key(week) <= week_sort_key(latest_uploaded_week)
    },
    key=week_sort_key,
  )

  values = []
  for week in available_weeks:
    activation_values = [
      activation_for_model_week(store, source_type, launch_year, model, week)
      for source_type, launch_year, model in model_scope
    ]
    valid_values = [value for value in activation_values if value is not None]
    if valid_values:
      total_activation = sum(valid_values)
      if total_activation > 0:
        values.append(int(round(total_activation)))
  return values[-max_points:]


def metric_row(result: dict, records: list[dict]):
  kpis = result["kpis"]
  cols = st.columns(4)
  trend_values = [row["count"] for row in result.get("trend", [])]
  act_trend_values = activation_trend_values(records, result.get("filters", {}))
  target_total_count = kpis.get("target_total_count", 0) or 0
  target_over_count = kpis.get("target_over_count", 0)
  target_accent = LINE_COLOR if target_total_count == 0 else ACCENT_GREEN if target_over_count == 0 else ACCENT_RED
  target_note = (
    f"{whole(kpis.get('target_hit_count'))} / {whole(target_total_count)} models on target"
    if target_total_count
    else "No comparable CFR target"
  )
  target_delta = (
    f"{whole(target_over_count)} models over target"
    if target_total_count
    else "Need CFR(A) for model + Target"
  )
  cards = [
    {
      "label": "Filtered CFR",
      "value": pct(kpis["filtered_cfr"]),
      "note": f"{whole(kpis['filtered_failure_qty'])} failures / {whole(kpis['derived_act'])} derived ACT",
      "accent": BAR_COLOR,
      "delta": f"Target CFR: {pct(kpis['target_cfr'])}",
      "spark": "",
    },
    {
      "label": "Failure Qty",
      "value": whole(kpis["filtered_failure_qty"]),
      "note": "Filtered raw-data failures",
      "accent": ACCENT_RED,
      "delta": f"Latest week: {whole(kpis['latest_count'])}",
      "spark": sparkline_svg(trend_values, ACCENT_RED),
    },
    {
      "label": "Derived ACT",
      "value": whole(kpis["derived_act"]),
      "note": f"From SUMMARY_IEC CFR rule; {whole(kpis['act_model_count'])} ACT models",
      "accent": LINE_COLOR,
      "delta": "ACT history trend" if len(act_trend_values) >= 2 else "Current-upload estimate only",
      "spark": sparkline_svg(act_trend_values, LINE_COLOR),
    },
    {
      "label": "Target Hit Rate",
      "value": pct(kpis.get("target_hit_rate")),
      "note": target_note,
      "accent": target_accent,
      "delta": target_delta,
      "spark": "",
    },
  ]
  for column, card in zip(cols, cards):
    column.markdown(
      f"""
      <div class="metric-card" style="--accent: {card['accent']};">
        <div class="metric-label">{card['label']}</div>
        <div class="metric-value">{card['value']}</div>
        <div class="metric-note">{card['note']}</div>
        <div class="metric-delta">{card['delta']}</div>
        {card['spark']}
      </div>
      """,
      unsafe_allow_html=True,
    )


def bar_chart(frame: pd.DataFrame, x_column: str, y_column: str, height: int = 260):
  axis_values = nice_axis_values(frame[y_column].max() if not frame.empty else 0)
  axis_domain = [0, axis_values[-1]]
  chart = (
    alt.Chart(frame)
    .mark_bar(color=BAR_COLOR, cornerRadiusTopLeft=3, cornerRadiusTopRight=3)
    .encode(
      y=alt.Y(f"{x_column}:N", sort="-x", axis=alt.Axis(title=None, labelLimit=180)),
      x=alt.X(
        f"{y_column}:Q",
        scale=alt.Scale(domain=axis_domain),
        axis=alt.Axis(title=None, values=axis_values),
      ),
      tooltip=[
        alt.Tooltip(f"{x_column}:N", title="Item"),
        alt.Tooltip(f"{y_column}:Q", title="Count"),
      ],
    )
    .properties(height=height)
  )
  st.altair_chart(chart, width="stretch")


def rows_to_pareto_frame(rows: list[dict], top_n: int = 10) -> pd.DataFrame:
  frame = pd.DataFrame(rows).copy()
  if frame.empty:
    return frame
  frame["_source_order"] = range(len(frame))
  frame = frame.sort_values(["count", "_source_order"], ascending=[False, True]).drop(columns=["_source_order"])
  if "share" not in frame:
    total = frame["count"].sum() or 1
    frame["share"] = frame["count"] / total * 100

  if len(frame) > top_n:
    top_rows = frame.head(top_n).copy()
    other_rows = frame.iloc[top_n:]
    others = {
      "label": "Others",
      "count": other_rows["count"].sum(),
      "share": other_rows["share"].sum(),
    }
    for column_name in other_rows.columns:
      if column_name in {"label", "count", "share", "cumulative", "cumulative_count"}:
        continue
      if pd.api.types.is_numeric_dtype(other_rows[column_name]):
        others[column_name] = other_rows[column_name].sum()
      else:
        others[column_name] = "Multiple"
    frame = pd.concat([top_rows, pd.DataFrame([others])], ignore_index=True)
  else:
    frame = frame.head(top_n).copy()

  frame["cumulative_count"] = frame["count"].cumsum()
  frame["cumulative"] = frame["share"].cumsum()
  if not frame.empty:
    frame.loc[frame.index[-1], "cumulative"] = 100.0
  frame["short_label"] = frame["label"].str.slice(0, 26)
  return frame


def render_pareto_frame(title: str, rows: list[dict], empty_message: str):
  st.subheader(title)
  pareto = rows_to_pareto_frame(rows)
  if pareto.empty:
    st.info(empty_message)
    return
  count_axis_values = nice_axis_values(pareto["count"].max())
  count_axis_domain = [0, count_axis_values[-1]]

  bars = (
    alt.Chart(pareto)
    .mark_bar(color=BAR_LIGHT_COLOR, cornerRadiusTopLeft=3, cornerRadiusTopRight=3)
    .encode(
      x=alt.X("short_label:N", sort=None, axis=alt.Axis(title=None, labelAngle=-35, labelLimit=90)),
      y=alt.Y(
        "count:Q",
        scale=alt.Scale(domain=count_axis_domain),
        axis=alt.Axis(title=None, values=count_axis_values, grid=True),
      ),
      tooltip=[
        alt.Tooltip("label:N", title="Item"),
        alt.Tooltip("count:Q", title="Failure Qty"),
        alt.Tooltip("share:Q", title="Share", format=".1f"),
        alt.Tooltip("cumulative:Q", title="Cumulative", format=".1f"),
      ],
    )
  )
  line = (
    alt.Chart(pareto)
    .mark_line(color=LINE_COLOR, strokeWidth=2.8)
    .encode(
      x=alt.X("short_label:N", sort=None),
      y=alt.Y(
        "cumulative:Q",
        scale=alt.Scale(domain=[0, 100]),
        axis=alt.Axis(
          title=None,
          orient="right",
          values=[0, 50, 100],
          labelExpr="datum.value + '%'",
          grid=False,
        ),
      ),
    )
  )
  points = (
    alt.Chart(pareto)
    .mark_point(color=LINE_COLOR, filled=True, size=62)
    .encode(
      x=alt.X("short_label:N", sort=None),
      y=alt.Y("cumulative:Q", scale=alt.Scale(domain=[0, 100]), axis=None),
      tooltip=[
        alt.Tooltip("label:N", title="Item"),
        alt.Tooltip("cumulative:Q", title="Cumulative", format=".1f"),
      ],
    )
  )
  st.altair_chart(
    alt.layer(bars, line, points).resolve_scale(y="independent").properties(height=320),
    width="stretch",
  )

  table = pareto.copy()
  table.insert(0, "rank", range(1, len(table) + 1))
  table["share"] = table["share"].map(lambda value: f"{value:.1f}%")
  table["cumulative"] = table["cumulative"].map(lambda value: f"{value:.1f}%")
  st.dataframe(
    table[["rank", "label", "count", "share", "cumulative"]],
    width="stretch",
    hide_index=True,
  )


def render_ratio(title: str, rows: list[dict]):
  st.subheader(title)
  if not rows:
    st.info("No records in the current filter.")
    return
  frame = pd.DataFrame(rows).head(10)
  bar_chart(frame, "label", "count", height=max(240, len(frame) * 28))
  table = frame.copy()
  table.insert(0, "rank", range(1, len(table) + 1))
  table["share"] = table["share"].map(lambda value: f"{value:.1f}%")
  table = table.rename(columns={"label": "item", "count": "failure_qty", "share": "share"})
  st.dataframe(
    table[["rank", "item", "failure_qty", "share"]],
    width="stretch",
    hide_index=True,
  )


def render_trend(result: dict):
  st.subheader("Weekly Failure Trend")
  trend = pd.DataFrame(result["trend"])
  if trend.empty:
    st.info("No trend data in the current filter.")
    return
  axis_values = nice_axis_values(trend["count"].max())
  axis_domain = [0, axis_values[-1]]
  chart = (
    alt.Chart(trend)
    .mark_line(point=True, color="#138a8e", strokeWidth=2.5)
    .encode(
      x=alt.X("week:N", axis=alt.Axis(labelAngle=-45, title=None)),
      y=alt.Y(
        "count:Q",
        scale=alt.Scale(domain=axis_domain),
        axis=alt.Axis(title=None, values=axis_values),
      ),
      tooltip=[
        alt.Tooltip("week:N", title="Week"),
        alt.Tooltip("count:Q", title="Failure Qty"),
      ],
    )
    .properties(height=300)
  )
  st.altair_chart(chart, width="stretch")


def selected_filter_values(value) -> set[str]:
  if value is None:
    return set()
  if isinstance(value, str):
    values = [value]
  else:
    values = list(value)
  return {str(item).strip() for item in values if str(item).strip()}


def record_filter_value(record: dict, column_name: str) -> str:
  text = str(record.get(column_name, "") or "").strip()
  return text if text else "(blank)"


def filtered_records_for_filters(records: list[dict], filters: dict[str, list[str]]) -> list[dict]:
  filtered = records
  for filter_key, column_name in FILTER_FIELDS.items():
    selected = selected_filter_values(filters.get(filter_key, []))
    if not selected:
      continue
    filtered = [
      record
      for record in filtered
      if record_filter_value(record, column_name) in selected
    ]
  return filtered


def act_scope_records_for_filters(records: list[dict], filters: dict[str, list[str]]) -> list[dict]:
  act_filters = {
    filter_key: filters.get(filter_key, [])
    for filter_key in ACT_SCOPE_FILTERS
  }
  return filtered_records_for_filters(records, act_filters)


def model_scope_for_filters(records: list[dict], filters: dict[str, list[str]]) -> list[tuple[str, str, str]]:
  scope_records = act_scope_records_for_filters(records, filters)
  return sorted(
    {
      (
        normalize_source_type(record.get("source_type", "")),
        launch_year_from_record(record),
        act_model_from_record(record),
      )
      for record in scope_records
      if launch_year_from_record(record) and act_model_from_record(record)
    }
  )


def activation_for_model_week(
  store: dict[tuple[str, str, str, str], float],
  source_type: str,
  launch_year: str,
  model: str,
  week: str,
) -> float | None:
  exact = store.get((source_type, launch_year, model, week))
  if exact is not None:
    return exact

  fallback_values = [
    value
    for (stored_source, stored_year, stored_model, stored_week), value in store.items()
    if stored_model == model and stored_week == week
    if stored_year in ("", launch_year)
    if stored_source == source_type or source_type == "Uploaded"
  ]
  if not fallback_values:
    return None
  return sum(fallback_values)


def cumulative_cfr_trend_data(records: list[dict], filters: dict[str, list[str]]) -> tuple[pd.DataFrame, str]:
  store = activation_history_store()
  if not store:
    return pd.DataFrame(), "No activation history is available yet."

  filtered_records = filtered_records_for_filters(records, filters)
  if not filtered_records:
    return pd.DataFrame(), "No failure records match the current filter selection."

  model_scope = model_scope_for_filters(records, filters)
  if not model_scope:
    return pd.DataFrame(), "No model scope is available for the current filter selection."

  failure_by_week = Counter(
    str(record.get("Week", "")).strip()
    for record in filtered_records
    if str(record.get("Week", "")).strip()
  )
  latest_uploaded_week = latest_week_from_records(records)
  available_weeks = sorted(
    {
      week
      for source_type, launch_year, model in model_scope
      for stored_source, stored_year, stored_model, week in store
      if stored_model == model and (stored_source == source_type or source_type == "Uploaded")
      if stored_year in ("", launch_year)
      if not latest_uploaded_week or week_sort_key(week) <= week_sort_key(latest_uploaded_week)
    },
    key=week_sort_key,
  )
  if not available_weeks:
    return pd.DataFrame(), "No activation weeks are available for the current filter selection."

  cumulative_activation_by_week = {}
  for week in available_weeks:
    activation_values = [
      activation_for_model_week(store, source_type, launch_year, model, week)
      for source_type, launch_year, model in model_scope
    ]
    valid_values = [value for value in activation_values if value is not None]
    cumulative_activation_by_week[week] = sum(valid_values) if valid_values else None

  trend_rows = []
  cumulative_failure = 0
  for current_week in available_weeks:
    weekly_failure = failure_by_week.get(current_week, 0)
    cumulative_failure += weekly_failure
    cumulative_activation = cumulative_activation_by_week.get(current_week)
    if cumulative_activation is None or cumulative_activation <= 0:
      continue

    trend_rows.append(
      {
        "end_week": current_week,
        "weekly_failure": weekly_failure,
        "cumulative_failure": cumulative_failure,
        "cumulative_activation": cumulative_activation,
        "cumulative_cfr": cumulative_failure / cumulative_activation,
      }
    )

  trend = pd.DataFrame(trend_rows)
  if trend.empty:
    return pd.DataFrame(), "No cumulative CFR can be calculated for the current filter selection."
  return trend, ""


def render_interval_cfr_trend(records: list[dict], filters: dict[str, list[str]]):
  st.subheader("Cumulative CFR Trend")
  trend, message = cumulative_cfr_trend_data(records, filters)
  if trend.empty:
    st.info(message)
    return

  chart = (
    alt.Chart(trend)
    .mark_line(point=True, color=LINE_COLOR, strokeWidth=2.8)
    .encode(
      x=alt.X("end_week:N", axis=alt.Axis(labelAngle=-35, title=None)),
      y=alt.Y(
        "cumulative_cfr:Q",
        axis=alt.Axis(title=None, format=".2%"),
      ),
      tooltip=[
        alt.Tooltip("end_week:N", title="Week"),
        alt.Tooltip("cumulative_cfr:Q", title="Cumulative CFR", format=".2%"),
        alt.Tooltip("cumulative_failure:Q", title="TTL Failures", format=",.0f"),
        alt.Tooltip("cumulative_activation:Q", title="Cumulative ACT", format=",.0f"),
      ],
    )
    .properties(height=300)
  )
  st.altair_chart(chart, width="stretch")


def reset_group_compare_if_upload_changed(upload_payloads: tuple[tuple[str, bytes], ...]):
  upload_signature = tuple((filename, len(file_bytes)) for filename, file_bytes in upload_payloads)
  if st.session_state.get("group_compare_upload_signature") == upload_signature:
    return
  st.session_state["group_compare_upload_signature"] = upload_signature
  st.session_state["group_compare_groups"] = []


def next_group_name(groups: list[dict]) -> str:
  return f"G{len(groups) + 1}"


def unique_group_name(label: str, groups: list[dict]) -> str:
  existing = {str(group.get("label", "")).strip() for group in groups}
  if label not in existing:
    return label
  suffix = 2
  while f"{label} ({suffix})" in existing:
    suffix += 1
  return f"{label} ({suffix})"


def group_latest_rows(trend_all: pd.DataFrame) -> list[dict]:
  rows = []
  for group, group_rows in trend_all.groupby("group", sort=False):
    sorted_rows = group_rows.sort_values("end_week", key=lambda series: series.map(week_sort_key))
    latest_row = sorted_rows.iloc[-1]
    previous_row = sorted_rows.iloc[-2] if len(sorted_rows) > 1 else None
    previous_cfr = previous_row["cumulative_cfr"] if previous_row is not None else None
    latest_cfr = latest_row["cumulative_cfr"]
    change = latest_cfr - previous_cfr if previous_cfr is not None else None
    relative_change = (
      change / previous_cfr
      if previous_cfr not in (None, 0)
      else None
    )
    alert = ""
    if relative_change is not None and relative_change >= 0.2:
      alert = "Sharp increase"
    elif previous_cfr in (None, 0) and latest_cfr > 0:
      alert = "New CFR"

    rows.append(
      {
        "group": group,
        "latest_week": latest_row["end_week"],
        "latest_cfr": latest_cfr,
        "previous_cfr": previous_cfr,
        "wow_change": change,
        "alert": alert or "OK",
        "cumulative_failure": latest_row["cumulative_failure"],
        "cumulative_activation": latest_row["cumulative_activation"],
      }
    )
  return rows


def pp(value: float | None) -> str:
  if value is None:
    return "N/A"
  sign = "+" if value > 0 else ""
  return f"{sign}{value * 100:.2f} pp"


def render_group_insight_cards(latest_rows: list[dict]):
  if not latest_rows:
    return

  highest = max(latest_rows, key=lambda row: row["latest_cfr"])
  lowest = min(latest_rows, key=lambda row: row["latest_cfr"])
  gap = highest["latest_cfr"] - lowest["latest_cfr"]
  latest_week = max((row["latest_week"] for row in latest_rows), key=week_sort_key)

  cards = [
    ("Highest CFR", highest["group"], pct(highest["latest_cfr"])),
    ("Lowest CFR", lowest["group"], pct(lowest["latest_cfr"])),
    ("Largest Gap", f"{highest['group']} vs {lowest['group']}", pp(gap)),
    ("Latest Week", latest_week, f"{len(latest_rows)} groups"),
  ]
  columns = st.columns(4)
  for column, (label, title, value) in zip(columns, cards):
    with column:
      st.markdown(
        f"""
        <div class="metric-card" style="--accent: {LINE_COLOR}; min-height: 118px;">
          <div class="metric-label">{html_escape(label)}</div>
          <div class="metric-value" style="font-size: 1.15rem;">{html_escape(title)}</div>
          <div class="metric-note">{html_escape(value)}</div>
        </div>
        """,
        unsafe_allow_html=True,
      )


def render_group_compare(records: list[dict]):
  st.markdown("### Group CFR Compare")
  st.caption("Cumulative CFR = TTL Failures / Cumulative ACT. TTL Failures is cumulative through each week for the selected group scope.")
  groups = st.session_state.setdefault("group_compare_groups", [])

  builder_col, group_col = st.columns([1, 1.25])
  with builder_col:
    st.markdown("#### Group Builder")
    draft_filters = filter_controls(records, "compare_filter")
    default_name = next_group_name(groups)
    group_name = st.text_input("Group name", value=default_name, key="compare_group_name")
    add_col, clear_col = st.columns([1, 1])
    with add_col:
      add_clicked = st.button("Add Group", type="primary", width="stretch")
    with clear_col:
      clear_clicked = st.button("Clear Groups", width="stretch", disabled=not groups)

    if clear_clicked:
      st.session_state["group_compare_groups"] = []
      st.rerun()

    if add_clicked:
      snapshot = clean_filter_snapshot(draft_filters)
      duplicate_label = matching_group_label(snapshot, groups)
      if duplicate_label:
        st.warning(f"This group has the same scope as {duplicate_label}.")
      else:
        label = unique_group_name(group_name.strip() or default_name, groups)
        st.session_state["group_compare_groups"].append(
          {
            "label": label,
            "filters": snapshot,
          }
        )
        st.rerun()

  with group_col:
    st.markdown("#### Groups")
    if not groups:
      st.info("Add at least one group to show CFR trend lines.")
    for index, group in enumerate(list(groups)):
      with st.container():
        columns = st.columns([1.15, 3.2, 0.95])
        with columns[0]:
          st.markdown(f"**{html_escape(group['label'])}**")
        with columns[1]:
          st.caption(filter_snapshot_summary(group.get("filters", {})))
        with columns[2]:
          if st.button("Remove", key=f"remove_compare_group_{index}", width="stretch"):
            del st.session_state["group_compare_groups"][index]
            st.rerun()
        with st.expander(f"{group['label']} details", expanded=False):
          st.dataframe(
            filter_snapshot_details(group.get("filters", {})),
            width="stretch",
            hide_index=True,
          )

  if not groups:
    return

  trend_frames = []
  skipped = []
  for index, group in enumerate(groups):
    trend, message = cumulative_cfr_trend_data(records, group.get("filters", {}))
    if trend.empty:
      skipped.append(f"{group['label']}: {message}")
      continue
    trend = trend.copy()
    trend["group"] = group["label"]
    trend["group_order"] = index
    trend_frames.append(trend)

  if not trend_frames:
    st.info("No cumulative CFR can be calculated for the added groups.")
    for message in skipped:
      st.caption(message)
    return

  trend_all = pd.concat(trend_frames, ignore_index=True)
  latest_rows = group_latest_rows(trend_all)
  render_group_insight_cards(latest_rows)

  chart_metric = st.radio(
    "Chart metric",
    ["Cumulative CFR", "Weekly Failure Qty"],
    horizontal=True,
    key="group_compare_chart_metric",
  )

  color_range = [
    GROUP_COMPARE_COLORS[index % len(GROUP_COMPARE_COLORS)]
    for index, _ in enumerate(groups)
  ]
  color_scale = alt.Scale(
    domain=[group["label"] for group in groups],
    range=color_range,
  )
  week_order = sorted(trend_all["end_week"].dropna().unique(), key=week_sort_key)
  if chart_metric == "Weekly Failure Qty":
    y_field = "weekly_failure:Q"
    y_axis = alt.Axis(title=None, format=",.0f")
    tooltips = [
      alt.Tooltip("group:N", title="Group"),
      alt.Tooltip("end_week:N", title="Week"),
      alt.Tooltip("weekly_failure:Q", title="Weekly Failures", format=",.0f"),
      alt.Tooltip("cumulative_failure:Q", title="TTL Failures", format=",.0f"),
    ]
  else:
    y_field = "cumulative_cfr:Q"
    y_axis = alt.Axis(title=None, format=".2%")
    tooltips = [
      alt.Tooltip("group:N", title="Group"),
      alt.Tooltip("end_week:N", title="Week"),
      alt.Tooltip("cumulative_cfr:Q", title="Cumulative CFR", format=".2%"),
      alt.Tooltip("cumulative_failure:Q", title="TTL Failures", format=",.0f"),
      alt.Tooltip("cumulative_activation:Q", title="Cumulative ACT", format=",.0f"),
    ]

  chart = (
    alt.Chart(trend_all)
    .mark_line(point=True, strokeWidth=2.8)
    .encode(
      x=alt.X("end_week:N", axis=alt.Axis(labelAngle=-35, title=None), sort=week_order),
      y=alt.Y(y_field, axis=y_axis),
      color=alt.Color("group:N", scale=color_scale, legend=alt.Legend(title="Group")),
      tooltip=tooltips,
    )
    .properties(height=360)
  )
  st.altair_chart(chart, width="stretch")

  latest_by_group = {row["group"]: row for row in latest_rows}
  best_cfr = min((row["latest_cfr"] for row in latest_rows), default=None)
  summary_rows = []
  for group in groups:
    latest_row = latest_by_group.get(group["label"])
    if not latest_row:
      continue
    summary_rows.append(
      {
        "Group": group["label"],
        "Latest Week": latest_row["latest_week"],
        "Latest CFR": pct(latest_row["latest_cfr"]),
        "Previous CFR": pct(latest_row["previous_cfr"]),
        "WoW Change": pp(latest_row["wow_change"]),
        "Gap vs Best": pp(latest_row["latest_cfr"] - best_cfr) if best_cfr is not None else "N/A",
        "Alert": latest_row["alert"],
        "TTL Failures": whole(latest_row["cumulative_failure"]),
        "Cumulative ACT": whole(latest_row["cumulative_activation"]),
        "Scope": filter_snapshot_summary(group.get("filters", {})),
      }
    )
  if summary_rows:
    st.dataframe(pd.DataFrame(summary_rows), width="stretch", hide_index=True)
  for message in skipped:
    st.caption(message)


def render_pareto(result: dict):
  render_pareto_frame(
    "PROBLEM_Mapping Pareto",
    result["pareto"],
    "No problem data in the current filter.",
  )


def render_action_insight(result: dict):
  rows = result.get("action_desc", [])
  st.subheader("ACTION_DESC Result")
  st.caption("Based on the current filter selections. ACTION_DESC is shown as a result insight, not as another filter.")

  if not rows:
    st.info("No ACTION_DESC data in the current filter.")
    return

  top_rows = rows[:3]
  card_columns = st.columns(3)
  for column, row in zip(card_columns, top_rows):
    column.markdown(
      f"""
      <div class="action-card">
        <div class="action-label">ACTION_DESC</div>
        <div class="action-title">{html_escape(row["label"])}</div>
        <div class="action-number">
          <strong>{whole(row["count"])}</strong>
          <span>{row["share"]:.1f}%</span>
        </div>
        <div class="action-note">Top Problem: {html_escape(row.get("top_problem", "N/A"))}</div>
      </div>
      """,
      unsafe_allow_html=True,
    )

  left, right = st.columns([1.05, 1])
  with left:
    render_pareto_frame(
      "ACTION_DESC Pareto",
      rows,
      "No ACTION_DESC data in the current filter.",
    )
  with right:
    detail = rows_to_pareto_frame(rows)
    if detail.empty:
      st.info("No ACTION_DESC detail data in the current filter.")
      return
    detail = detail.rename(
      columns={
        "label": "action_desc",
        "count": "failure_qty",
        "share": "share",
        "cumulative": "cumulative",
        "top_model": "top_model",
        "top_problem": "top_problem",
        "latest_week_count": "latest_week",
      }
    )
    detail.insert(0, "rank", range(1, len(detail) + 1))
    detail["share"] = detail["share"].map(lambda value: f"{value:.1f}%")
    detail["cumulative"] = detail["cumulative"].map(lambda value: f"{value:.1f}%")
    st.subheader("ACTION_DESC Detail")
    st.dataframe(
      detail[["rank", "action_desc", "failure_qty", "share", "cumulative", "top_model", "top_problem", "latest_week"]],
      width="stretch",
      hide_index=True,
    )


def render_module_pareto(result: dict):
  render_pareto_frame(
    "MUC_MODULE Pareto",
    result["module_ratio"],
    "No module data in the current filter.",
  )


def render_heatmap(result: dict):
  st.subheader("Watch Matrix Heatmap")
  st.caption("Each cell shows failure count with CFR in parentheses when ACT is available.")
  columns = result["heatmap_columns"]
  rows = result["heatmap"]
  if not columns or not rows:
    st.info("No matrix data in the current filter.")
    return

  count_frame = pd.DataFrame(
    [[cell["count"] for cell in row["values"]] for row in rows],
    index=[row["label"] for row in rows],
    columns=columns,
  )
  cfr_frame = pd.DataFrame(
    [[cell["cfr"] for cell in row["values"]] for row in rows],
    index=[row["label"] for row in rows],
    columns=columns,
  )

  raw_rows = []
  has_cfr = cfr_frame.notna().any().any()
  for row_label in count_frame.index:
    for column_label in count_frame.columns:
      count = int(count_frame.loc[row_label, column_label])
      cfr_value = cfr_frame.loc[row_label, column_label]
      color_value = cfr_value * 100 if has_cfr and pd.notna(cfr_value) else count
      raw_rows.append(
        {
          "model": row_label,
          "problem": column_label,
          "count": count,
          "cfr_pct": cfr_value * 100 if pd.notna(cfr_value) else None,
          "color_value": color_value,
          "count_label": str(count) if count else "-",
          "cfr_label": f"CFR {cfr_value * 100:.2f}%" if count and pd.notna(cfr_value) else "",
        }
      )

  frame = pd.DataFrame(raw_rows)
  rect = (
    alt.Chart(frame)
    .mark_rect(stroke="#ffffff", strokeWidth=1.5)
    .encode(
      x=alt.X("problem:N", sort=columns, axis=alt.Axis(title=None, labelAngle=-30, labelLimit=120)),
      y=alt.Y("model:N", sort=[row["label"] for row in rows], axis=alt.Axis(title=None, labelLimit=120)),
      color=alt.Color(
        "color_value:Q",
        scale=alt.Scale(range=["#e9f7f5", "#a8dad7", "#5fb7bc", "#1a9295"]),
        legend=alt.Legend(title="CFR %" if has_cfr else "Failure Qty"),
      ),
      tooltip=[
        alt.Tooltip("model:N", title="Model"),
        alt.Tooltip("problem:N", title="Problem"),
        alt.Tooltip("count:Q", title="Failure Qty"),
        alt.Tooltip("cfr_pct:Q", title="CFR %", format=".2f"),
      ],
    )
  )
  count_text = (
    alt.Chart(frame)
    .mark_text(
      baseline="middle",
      dy=-8,
      fontSize=13,
      fontWeight="bold",
    )
    .encode(
      x=alt.X("problem:N", sort=columns),
      y=alt.Y("model:N", sort=[row["label"] for row in rows]),
      text="count_label:N",
      opacity=alt.condition(alt.datum.count > 0, alt.value(1), alt.value(0.65)),
    )
  )
  cfr_text = (
    alt.Chart(frame)
    .mark_text(
      baseline="middle",
      dy=9,
      fontSize=10,
      fontWeight="bold",
      color="#335a61",
    )
    .encode(
      x=alt.X("problem:N", sort=columns),
      y=alt.Y("model:N", sort=[row["label"] for row in rows]),
      text="cfr_label:N",
      opacity=alt.condition(alt.datum.count > 0, alt.value(1), alt.value(0)),
    )
  )
  st.altair_chart(
    alt.layer(rect, count_text, cfr_text).properties(height=max(340, len(rows) * 46)),
    width="stretch",
  )


def render_file_summary(parsed: dict):
  files = parsed.get("files", [])
  if not files:
    return
  with st.expander("Loaded weekly raw data", expanded=False):
    for file_info in files:
      st.markdown(
        f"- **{file_info['source_type']}** - `{file_info['filename']}` - {file_info['rows']} rows"
      )


def render_parse_diagnostics(parsed: dict):
  missing_columns = parsed.get("missing_columns", {})
  files = parsed.get("files", [])
  if files:
    with st.expander("Workbook parsing details", expanded=True):
      for file_info in files:
        st.markdown(
          f"- **{file_info['filename']}**: {file_info['rows']} raw-data rows, "
          f"{len(file_info.get('sheets', []))} sheets"
        )
  if missing_columns:
    st.error("Some uploaded workbooks are missing the required sheet or columns.")
    with st.expander("Missing sheet / column details", expanded=True):
      for filename, columns in missing_columns.items():
        st.markdown(f"- `{filename}`: {', '.join(columns)}")


def render_change_log():
  entries = [
    ("2026-06-11", "CFR Watch Board initial release: uploaded Gaming / PC NB raw data and added KPI, trend, Pareto, heatmap, and top item views."),
    ("2026-06-12", "Streamlit POC visual tuning: improved filters, Pareto, heatmap colors, and chart labels."),
    ("2026-06-15", "Parser hardening: improved upload feedback, supported 2025 CFR workbook data, and added ACT summary model prefix matching."),
    ("2026-06-26", "Added ACTION_DESC result insights with top cards, Pareto, and detail table linked to current filters."),
    ("2026-07-09", "Added ACT seed history and Cumulative CFR Trend, with week labels simplified on the time axis."),
    ("2026-07-10", "Added automatic ACT extraction after upload, GitHub-backed activation_history.csv updates, and CSV download fallback."),
    ("2026-07-11", "Added Target Hit Rate KPI using SUMMARY_IEC CFR(A) for model versus Target, while avoiding Series CFR(A) Average."),
    ("2026-07-13", "Added Group CFR Compare mode with G1/G2 scopes, duplicate-group checks, WoW/Gap/Alert fields, and chart metric switching."),
    ("2026-08-04", "Added ACT table persistence: 2025 ACT uses MODEL_GROUP, 2026 ACT uses ORG_MODEL(PRODUCT_DESC), existing ACT table values are preserved, and the Site Change Log encoding was repaired."),
    ("2026-08-21", "Backfilled ACT table history for W2631-W2633 from uploaded weekly raw data, added ACT Persistence Guard, and refreshed ACT history cache whenever the ACT table changes."),
    ("2026-08-21", "Made Group CFR Compare and Cumulative CFR Trend year-aware: uploaded rows use launch year to select the matching 2025 ACT or 2026 ACT table values."),
    ("2026-08-21", "Added Launch Year as a selectable filter in Overview Dashboard and Group CFR Compare so groups can be built separately for 2025 ACT and 2026 ACT scopes."),
  ]
  items = "\n".join(
    f"<li><strong>{html_escape(date)}</strong> {html_escape(message)}</li>"
    for date, message in entries
  )
  st.markdown(
    f"""
    <div class="change-log">
      <h3>Site Change Log</h3>
      <ol>
        {items}
      </ol>
    </div>
    """,
    unsafe_allow_html=True,
  )


def main():
  reset_session_if_app_version_changed()
  apply_page_style()
  st.markdown('<div class="cfr-title">CFR Watch Board</div>', unsafe_allow_html=True)
  st.markdown(
    '<div class="cfr-subtitle">Upload weekly Gaming NB and PC NB CFR workbooks, then filter by model, segment, ODM/OEM, module, and problem.</div>',
    unsafe_allow_html=True,
  )

  if not password_gate():
    return

  with st.sidebar:
    if st.button("Sign out"):
      st.session_state.pop("authenticated", None)
      st.rerun()
    render_act_persistence_preflight()

  uploaded_files = st.file_uploader(
    "Weekly CFR workbooks",
    type=["xlsx", "xlsm", "xls"],
    accept_multiple_files=True,
    help="Upload the weekly Gaming NB and PC NB files together.",
  )

  if not uploaded_files:
    st.info("Upload weekly CFR raw-data workbooks to start.")
    st.divider()
    render_change_log()
    return

  upload_payloads = build_upload_payloads(uploaded_files)
  reset_group_compare_if_upload_changed(upload_payloads)
  total_upload_mb = uploaded_size_mb(upload_payloads)
  parse_notice = st.empty()
  parse_notice.info(f"{upload_summary_text(upload_payloads)} Reading the Excel raw data now.")
  filter_wait_notice = st.sidebar.empty()
  filter_wait_notice.info("Filters will appear after the uploaded workbook is parsed.")
  if total_upload_mb > 80:
    st.warning(
      f"Uploaded files total {total_upload_mb:.1f} MB. Large Excel workbooks may take longer on Streamlit Community Cloud."
    )

  try:
    with st.spinner("Reading weekly CFR workbooks..."):
      parsed = parse_uploaded_payloads(upload_payloads, PARSE_CACHE_VERSION)
  except Exception as exc:
    st.error(f"CFR analysis failed: {exc}")
    return

  parse_notice.empty()
  filter_wait_notice.empty()
  records = parsed.get("records", [])
  if not records:
    st.warning("No raw-data records were found. Please confirm the workbook includes a 'raw data' sheet.")
    render_parse_diagnostics(parsed)
    return

  activation_history_result = remember_activation_snapshot(parsed)
  act_table_result = remember_act_table_snapshot(parsed)
  render_file_summary(parsed)
  render_act_table_status(act_table_result)
  render_activation_history_status(activation_history_result)

  view_mode = render_mode_selector()

  if view_mode == "Group Compare":
    render_group_compare(records)
    st.divider()
    render_change_log()
    return

  selections = selected_filters(records)
  result = analyze_dataset(
    records,
    summary_by_model=parsed.get("summary_by_model", {}),
    primary_dimension="org_model",
    breakdown_dimension="problem_mapping",
    filters=selections,
  )

  dashboard_header(result)
  metric_row(result, records)
  st.divider()

  render_interval_cfr_trend(records, selections)

  st.divider()
  render_trend(result)

  left, right = st.columns(2)
  with left:
    render_ratio("ORG_MODEL(PRODUCT_DESC) Share", result["share"])
  with right:
    render_module_pareto(result)

  left, right = st.columns([1, 1.2])
  with left:
    render_pareto(result)
  with right:
    render_heatmap(result)

  st.divider()
  render_action_insight(result)

  st.divider()
  render_change_log()


if __name__ == "__main__":
  main()





